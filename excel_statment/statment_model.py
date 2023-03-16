from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import os
import pyexcel as p
import pickle
import xlrd
import shutil

from excel_statment.properties_model import PhysicalProperties, MechanicalProperties, PropertiesDict, ConsolidationProperties, CyclicProperties
from descriptors import DataTypeValidation
from general.general_functions import create_path
from excel_statment.position_configs import PhysicalPropertyPosition, MechanicalPropertyPosition, c_fi_E_PropertyPosition, \
    DynamicsPropertyPosition, IdentificationColumns, GeneralDataColumns
from excel_statment.functions import str_df, float_df
from general.general_functions import read_json_file, create_json_file
import shelve
from loggers.logger import app_logger


class StatmentData:
    """Класс, хранящий данные ведомости"""
    object_name = DataTypeValidation(str)
    customer = DataTypeValidation(str)
    start_date = DataTypeValidation(datetime)
    end_date = DataTypeValidation(datetime)
    accreditation = DataTypeValidation(str)
    accreditation_key = DataTypeValidation(str)
    object_number = DataTypeValidation(str)
    shipment_number = DataTypeValidation(str)
    path = DataTypeValidation(str)

    def __init__(self, statment_path):
        if statment_path.endswith("xlsx"):
            wb = load_workbook(statment_path, data_only=True)
            object_name = str(wb["Лист1"][GeneralDataColumns["object_name"][0]].value)
            customer = str(wb["Лист1"][GeneralDataColumns["customer"][0]].value)
            accreditation = str(wb["Лист1"][GeneralDataColumns["accreditation"][0]].value)
            object_number = str(wb["Лист1"][GeneralDataColumns["object_number"][0]].value)
            start_date = wb["Лист1"][GeneralDataColumns["start_date"][0]].value
            end_date = wb["Лист1"][GeneralDataColumns["end_date"][0]].value
            shipment_number = wb["Лист1"][GeneralDataColumns["shipment_number"][0]].value
            wb.close()
        else:
            wb = xlrd.open_workbook(statment_path, formatting_info=True)
            sheet = wb.sheet_by_index(0)
            object_name = str(sheet.cell(*GeneralDataColumns["object_name"][1]).value)
            customer = str(sheet.cell(*GeneralDataColumns["customer"][1]).value)
            accreditation = str(sheet.cell(*GeneralDataColumns["accreditation"][1]).value)
            object_number = str(sheet.cell(*GeneralDataColumns["object_number"][1]).value)
            start_date = datetime(*xlrd.xldate_as_tuple(sheet.cell(*GeneralDataColumns["start_date"][1]).value, wb.datemode))
            end_date = datetime(*xlrd.xldate_as_tuple(sheet.cell(*GeneralDataColumns["end_date"][1]).value, wb.datemode))
            shipment_number = sheet.cell(*GeneralDataColumns["shipment_number"][1]).value
            #print("dfg", shipment_number)

        if accreditation in ["OAO", "ОАО"]:
            accreditation = "АО"
        elif accreditation == "OOO":
            accreditation = "ООО"

        assert not isinstance(self.start_date, datetime), "Не установлена дата начала опытов"
        assert not isinstance(self.end_date, datetime), "Не установлена дата окончания опытов"
        assert object_name, "Не указано имя объекта"
        assert customer, "Не указан заказчик"
        assert customer, "Не указан заказчик"
        assert accreditation, "Не указана аккредитация"
        if shipment_number is None:
            shipment_number = ""
        else:
            shipment_number = str(shipment_number)

        if accreditation in ["AO", "АО"]:
            self.accreditation_key = "новая"
        elif accreditation == "ООО" or accreditation == "OOO":
            self.accreditation_key = "2"

        self.object_name = object_name
        self.customer = customer
        self.accreditation = accreditation
        self.object_number = object_number
        self.start_date = start_date
        self.end_date = end_date
        self.shipment_number = shipment_number
        self.path = statment_path

    def __repr__(self):
        return str(self.__dict__)

    def get_shipment_number(self):

        if self.shipment_number:
            shipment_number = f" - {self.shipment_number}"
        else:
            shipment_number = ""

        return shipment_number

    def get_json(self):
        return self.__dict__

class GeneralParameters:
    """Класс, хранящий общие свойства по испытаниям"""
    def __init__(self, data: dict):
        for key in data:
            setattr(self, key, data[key])
        if not hasattr(self, "test_mode"):
            setattr(self, "test_mode", None)

    def __repr__(self):
        return str(self.__dict__)

    def get_json(self):
        return self.__dict__

class SaveDir:
    """Класс создает интерфейс для сохранения отчетов.
    Сигнал с директорией файла ведомости передается из класса открытия,
     после чего в этой директории создаются соответствующие папки.
     Название папки отчета передается в класс через коструктор mode"""

    def __init__(self, path=""):
        super().__init__()

        self._save_directory = "C:/"

        self.additional_dirs = []

        self.postfix = ""
        self.mode = ""

    def __getattr__(self, item):
        if item in self.additional_dirs:
            return self._save_directory + f"/{item}/"
        else:
            raise AttributeError

    @property
    def report_directory(self):
        return self._save_directory + f"/{self.mode}{self.postfix}/"

    @property
    def save_directory(self):
        return self._save_directory

    @property
    def arhive_directory(self):
        return self._save_directory + f"/Архив {self.mode}{self.postfix}/"

    @property
    def cvi_directory(self):
        return self._save_directory + f"/ЦВИ{self.postfix}/"

    @property
    def directory(self):
        return self._save_directory

    def _create_paths(self):
        create_path(self._save_directory)

        for path in [self.report_directory, self.arhive_directory, self.cvi_directory]:
            create_path(path)

        for path in [getattr(self, dir) for dir in self.additional_dirs]:
            create_path(path)

    def _create_save_directory(self, path, mode=""):
        """Создание папки и подпапок для сохранения отчета"""
        self._save_directory = path + "/" + mode

        self._create_paths()

        app_logger.info(f"Папка сохранения опытов {self._save_directory}")

    def check_dirs(self):
        self._create_paths()

    def clear_dirs(self):
        for path in [self.report_directory, self.arhive_directory, self.cvi_directory]:
            if os.path.exists(path):
                shutil.rmtree(path)

        for path in [getattr(self, dir) for dir in self.additional_dirs]:
            if os.path.exists(path):
                shutil.rmtree(path)

        self.check_dirs()

    def set_directory(self, dir, mode, postfix="", additional_dirs: list = []):
        """Получение пути к файлу ведомости excel"""
        self.additional_dirs = additional_dirs

        self.mode = mode

        if postfix:
            self.postfix = f" - {postfix}"
        else:
            self.postfix = ""

        self._create_save_directory(os.path.split(dir)[0], mode)


class Test:
    physical_properties = DataTypeValidation(PhysicalProperties)
    mechanical_properties = DataTypeValidation(MechanicalProperties, ConsolidationProperties)

    def __init__(self, test_class=None, data_frame=None, i=None, test_mode=None, K0_mode=None, sigma3_lim=None,
                 identification_column=None, physical_properties_dict=None, mechanical_properties_dict=None,
                 equipment=None, phi_mode=None):
        if physical_properties_dict and mechanical_properties_dict:
            self.physical_properties = PhysicalProperties()
            for attr in physical_properties_dict:
                if attr == "date":
                    if isinstance(physical_properties_dict[attr], datetime)  is None:
                        setattr(self.physical_properties, attr,
                                datetime.strptime(physical_properties_dict[attr].split(".")[0], "%Y-%m-%d %H:%M:%S"))
                        setattr(self.physical_properties, attr, physical_properties_dict[attr])
                    else:
                        setattr(self.physical_properties, attr, None)
                else:
                    setattr(self.physical_properties, attr, physical_properties_dict[attr])

            self.mechanical_properties = test_class()
            for attr in mechanical_properties_dict:
                setattr(self.mechanical_properties, attr, mechanical_properties_dict[attr])

        else:
            self.physical_properties = PhysicalProperties()

            sample_size = None
            if equipment == "АСИС ГТ.2.0.5 (150х300)":
                sample_size = (150, 300)

            self.physical_properties.defineProperties(data_frame=data_frame, string=i,
                                                      identification_column=identification_column,
                                                      sample_size=sample_size)
            self.mechanical_properties = test_class()

            if sigma3_lim:
                if phi_mode:
                    self.mechanical_properties.defineProperties(self.physical_properties, data_frame=data_frame,
                                                                string=i,
                                                                test_mode=test_mode, K0_mode=K0_mode,
                                                                sigma3_lim=sigma3_lim,
                                                                phi_mode=phi_mode)
                else:
                    self.mechanical_properties.defineProperties(self.physical_properties, data_frame=data_frame,
                                                                string=i,
                                                                test_mode=test_mode, K0_mode=K0_mode,
                                                                sigma3_lim=sigma3_lim)
            else:
                if phi_mode:
                    self.mechanical_properties.defineProperties(self.physical_properties, data_frame=data_frame,
                                                                string=i,
                                                                test_mode=test_mode, K0_mode=K0_mode, phi_mode=phi_mode)
                else:
                    self.mechanical_properties.defineProperties(self.physical_properties, data_frame=data_frame, string=i,
                                                            test_mode=test_mode, K0_mode=K0_mode)

    def get_json(self):
        return {
            "physical_properties": self.physical_properties.__dict__,
            "mechanical_properties": self.mechanical_properties.__dict__
        }

    def __repr__(self):
        return f"Физические свойства: {self.physical_properties}, Механические свойства {self.mechanical_properties}"

class Statment:
    tests = DataTypeValidation(dict)
    general_data = DataTypeValidation(StatmentData)
    general_parameters = DataTypeValidation(GeneralParameters)
    test_class = None
    current_test: str
    original_keys: list = None
    save_dir = DataTypeValidation(SaveDir)
    backup_dir = "Z:/DigitRock Models Backup"

    def __new__(cls):
        if not hasattr(cls, 'instance'):
            cls.instance = super(Statment, cls).__new__(cls)
        return cls.instance

    def __init__(self):
        self.test_class = None
        self.general_parameters = None
        self.tests = {}
        self.current_test = None
        self.save_dir = SaveDir()

    def setTestClass(self, cls):
        self.test_class = cls

    def setCurrentTest(self, lab):
        self.current_test = lab

    def getLaboratoryNumber(self):
        return self.tests[self.current_test].physical_properties.lab_number

    def readExcelFile(self, excel_path, identification_column):
        self.tests = {}

        assert self.general_parameters, "Определите начальные параметры"

        self.general_data = StatmentData(excel_path)

        data_frame = Statment.createDataFrame(excel_path)

        if not hasattr(self.general_parameters, "K0_mode"):
            self.general_parameters.K0_mode = None

        if not hasattr(self.general_parameters, "sigma3_lim"):
            self.general_parameters.sigma3_lim = None

        if not hasattr(self.general_parameters, "equipment"):
            self.general_parameters.equipment = None

        if not  hasattr(self.general_parameters, "phi_mode"):
            self.general_parameters.phi_mode = None

        if data_frame is not None:
            for i in range(len(data_frame["Лаб. № пробы"])):
                self.tests[str_df(data_frame.iat[i, PhysicalPropertyPosition["laboratory_number"][1]])] = \
                    Test(self.test_class, data_frame, i, self.general_parameters.test_mode,
                         self.general_parameters.K0_mode, self.general_parameters.sigma3_lim, identification_column,
                         equipment=self.general_parameters.equipment, phi_mode=self.general_parameters.phi_mode)
        self.original_keys = list(self.tests.keys())

    def setGeneralParameters(self, data):
        self.general_parameters = GeneralParameters(data)

    def sort(self, key):
        if key == "origin":
            self.tests = {key: self.tests[key] for key in self.original_keys}
        elif hasattr(self.tests[self.original_keys[0]].physical_properties, key):
            self.tests = dict(sorted(self.tests.items(), key=lambda x: getattr(self.tests[x[0]].physical_properties, key)))
        elif hasattr(self.tests[self.original_keys[0]].mechanical_properties, key):
            self.tests = dict(sorted(self.tests.items(), key=lambda x: getattr(self.tests[x[0]].mechanical_properties, key)))

    def dump(self, directory, name="statment.pickle"):
        general_data = {
            "tests": self.tests,
            "general_data": self.general_data,
            "general_parameters": self.general_parameters,
            "test_class": self.test_class,
            "original_keys": self.original_keys
        }
        with open(directory + "/" + name, "wb") as file:
            pickle.dump(general_data, file)

    def load(self, file):
        with open(file, 'rb') as f:
            data = pickle.load(f)
            self.tests = data["tests"]
            self.general_data = data["general_data"]
            self.general_parameters = data["general_parameters"]
            self.test_class = data["test_class"]
            self.original_keys = data["original_keys"]

    def save(self, models: list, models_names: list):
        date_format = "%m.%d.%Y %H-%M-%S"
        str_datetime = datetime.now().strftime(date_format)

        backup_object = os.path.join(self.backup_dir, self.general_data.object_number)
        backup_test = os.path.join(backup_object, self.general_parameters.test_mode)
        backup_date_path = os.path.join(backup_test, str_datetime)

        create_path(backup_object)
        create_path(backup_test)
        create_path(backup_date_path)

        paths = []
        for entry in os.scandir(backup_test):
            if entry.is_dir():
                paths.append(datetime.strptime(os.path.split(entry)[-1], date_format))

        if len(paths) > 3:
            shutil.rmtree(os.path.join(backup_test, min(paths).strftime(date_format)))

        shutil.copy(self.general_data.path, os.path.join(backup_date_path, os.path.split(self.general_data.path)[-1]))

        for model, model_name in zip(models, models_names):
            model_path = os.path.join(backup_date_path, self.general_parameters.test_mode)
            create_path(model_path)
            model.dump(os.path.join(model_path, model_name))

    @staticmethod
    def createDataFrame(excel_path) -> pd.DataFrame:
        """Функция считывания файла excel в датафрейм"""

        def resave_xls_to_xlsx(file):
            """Пересохраняет файл excel из формата xls в xlsx
                    Вернет имя нового файла
                    Если файл уже есть, то вернет его"""

            current_file_name = ""

            if file != "":
                # Проверим наличие документа Exel. Если он в старом формате то пересохраним в новый
                if file[-1] == "x":
                    current_file_name = file
                elif file[-1] == "s":
                    p3 = file + "x"
                    if os.path.exists(p3):
                        current_file_name = file
                        pass
                    else:
                        p.save_book_as(file_name=file,
                                       dest_file_name=p3)  # проверяем есть ли xlsx. Если нет то создаем копию файла в этом формате
                    file = p3
                    current_file_name = file
                else:
                    pass
                current_file_name = file

            return current_file_name

        '''if (excel_path.endswith("xlsx") or excel_path.endswith("xls")) and not read_xls:
            wb = resave_xls_to_xlsx(excel_path)
        elif excel_path.endswith("xls") and read_xls:
            wb = excel_path
        else:
            return None'''
        wb = excel_path
        df = pd.read_excel(wb, usecols="A:IV", skiprows=[0, 1, 3, 4, 5])
        df = df[df['Лаб. № пробы'].notna()]

        return df

    def __iter__(self):
        for key in self.tests:
            yield key

    def __getitem__(self, key):
        if not key in list(self.tests.keys()):
            return KeyError(f"No test with key {key}")
        return self.tests[key]

    def __len__(self):
        return len(self.tests)

    def __str__(self):
        customer = f"Общие данные:\n {self.general_data}\n"
        general_parameters = f"Данные опыта:\n {self.general_parameters}\n"
        tests = "Опыты\n" + "\n".join(f"{key}: {self.tests[key]}" for key in self.tests.keys())
        return customer + general_parameters + tests


if __name__ == '__main__':
    from excel_statment.properties_model import RCProperties
    s = Statment()
    s.setTestClass(MechanicalProperties)
    s.setGeneralParameters({'test_mode': "Трёхосное сжатие (F, C, E)", 'K0_mode': 'K0: По ГОСТ-65353'})
    s.readExcelFile(r"C:\Users\Пользователь\Desktop\test\818-20 Атомфлот - мех.xlsx", None)
    print(s)
