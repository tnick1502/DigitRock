from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from excel_statment.position_configs import GeneralDataColumns
import xlrd
import xlutils
import xlwt

def float_df(x):
    if str(x) != "nan" and str(x) != "NaT":
        try:
            return float(x)
        except ValueError:
            return x
        except TypeError:
            return x
    else:
        return None

def str_df(x):
    if str(x) != "nan" and str(x) != "NaT":
        return str(x)
    else:
        return None

def date_df(x):
    if str(x) == "nan" or str(x) == "NaT":
        return None
    else:
        return x.to_pydatetime()


def column_fullness_test(path, columns=[], initial_columns=[]):
    """Проверка заполнения заданной ячейки
    входные параметры:
        columns - список столбцов, которые проверяются на заполненность
        initial_columns - список столбцов, которые будут указателями, какие строки для columns проверять"""

    def test_one_columns_xlsx(wb, column, initial_columns):
        """Проверка заполнения одного столбца, с учетом initial"""
        all_is_okey = True
        initial_columns.append(("A", 0))
        for i in range(7, len(wb['Лист1']['A'])):
            if "None" not in [str(wb["Лист1"][x[0] + str(i)].value) for x in initial_columns]:
                if str(wb["Лист1"][column + str(i)].value) == "None":
                    all_is_okey = False
        return all_is_okey

    def test_one_columns_xls(wb, column, initial_columns):
        """Проверка заполнения одного столбца, с учетом initial"""
        sheet = wb.sheet_by_index(0)
        all_is_okey = True
        initial_columns.append(("A", 0))
        """print(sheet.nrows)
        for i in range(7, sheet.nrows):
            if str(sheet.cell(i, 0).value) == "None" or str(sheet.cell(i, 0).value) == "":
                len_wb = i
                break"""

        for i in range(7, sheet.nrows):
            if all([str(sheet.cell(i, x[1]).value) for x in initial_columns]):
                if not str(sheet.cell(i, column).value):
                    all_is_okey = False
        return all_is_okey

    if path.endswith("xlsx"):
        wb = load_workbook(path, data_only=True)
        for column in columns:
            all_is_okey = test_one_columns_xlsx(wb, column[0], initial_columns)
            if all_is_okey == False:
                return False
        wb.close()
    elif path.endswith("xls"):
        wb = xlrd.open_workbook(path, formatting_info=True)
        for column in columns:
            all_is_okey = test_one_columns_xls(wb, column[1], initial_columns)
            if all_is_okey == False:
                return False

    return True

def k0_test_type_column(test_type):
    """Функция возвращает столбцы для считывания K0 по типу испытания"""
    if test_type == "K0: K0nc из ведомости":
        return [["GZ", 207]]
    elif test_type == "K0: K0 из ведомости":
        return [["GY", 206]]
    else:
        return [["A", 0]]

def read_general_prameters(path):
    """Чтение данных заказчика, даты
        Передается документ excel, возвращает маркер False и данные, либо маркер True и имя ошибки"""

    if path.endswith("xlsx"):
        wb = load_workbook(path, data_only=True)
        data = {
            "object_name": str(wb["Лист1"][GeneralDataColumns["object_name"][0]].value),
            "customer": str(wb["Лист1"][GeneralDataColumns["customer"][0]].value),
            "accreditation": str(wb["Лист1"][GeneralDataColumns["accreditation"][0]].value),
            "object_number": str(wb["Лист1"][GeneralDataColumns["object_number"][0]].value),
            "start_date": wb["Лист1"][GeneralDataColumns["start_date"][0]].value,
            "end_date": wb["Лист1"][GeneralDataColumns["end_date"][0]].value,
        }
        wb.close()
    else:
        wb = xlrd.open_workbook(path, formatting_info=True)
        sheet = wb.sheet_by_index(0)
        data = {
            "object_name": str(sheet.cell(*GeneralDataColumns["object_name"][1]).value),
            "customer": str(sheet.cell(*GeneralDataColumns["customer"][1]).value),
            "accreditation": str(sheet.cell(*GeneralDataColumns["accreditation"][1]).value),
            "object_number": str(sheet.cell(*GeneralDataColumns["object_number"][1]).value),
            "start_date": sheet.cell(*GeneralDataColumns["start_date"][1]).value,
            "end_date": sheet.cell(*GeneralDataColumns["end_date"][1]).value,
        }

        try:
            data["start_date"] = datetime(*xlrd.xldate_as_tuple(data["start_date"], wb.datemode))
            data["end_date"] = datetime(*xlrd.xldate_as_tuple(data["end_date"], wb.datemode))
        except:
            pass

    for i in data:
        if data[i] == "None":
            return True, i

    if not isinstance(data["end_date"], datetime):
        return True, "Дата окончания опытов"

    if not isinstance(data["start_date"], datetime):
        return True, "Дата начала опытов"

    return False, data

def write_to_excel(path, Lab, params):
    """Запись в файл excel"""
    wb = load_workbook(path)

    wb["Лист1"]["HY5"] = "Сигма1, кПа"
    wb["Лист1"]["HZ5"] = "Сигма3, кПа"
    wb["Лист1"]["IA5"] = "Тау, кПа"
    wb["Лист1"]["IB5"] = "K0"
    wb["Лист1"]["IC5"] = "Частота, Гц"
    wb["Лист1"]["ID5"] = "Цикл разрушения"

    iLab_new = 0
    iLab = 0

    for i in range(7, len(wb['Лист1']['A']) + 5):

        if str(wb["Лист1"]['IG' + str(i)].value) == Lab:
            iLab_new = i
        elif str(wb["Лист1"]['A' + str(i)].value) == Lab:
            iLab = i

    if iLab_new != 0:
        cells = ["HW", "HX", "HY", "HZ", "IA", "IB", "IC", "ID", "IE", "IF"]
        for param, cell in zip(params, cells):
            wb["Лист1"][cell + str(iLab_new)] = param

    elif iLab != 0:
        cells = ["HW", "HX", "HY", "HZ", "IA", "IB", "IC", "ID", "IE", "IF"]
        for param, cell in zip(params, cells):
            wb["Лист1"][cell + str(iLab)] = param

    wb.save(path)

def set_cell_data(path: str, cell: str, value, sheet: str="Лист1", color=None)->None:
    """Запись в файл excel

    :argument path: путь к файлу excel в формате xlsx
    :argument cell: Ячейка ('A1', (0, 0))
    :argument value: Записываемое значение
    :argument sheet: Лист для записи
    :argument color: цвет шрифта записи

    :return None"""
    if path.endswith("xlsx"):
        wb = load_workbook(path)
        wb[sheet][cell[0]] = value
        if color:
            cell = wb[sheet][cell[0]]
            cell.font = Font(color=color)
        wb.save(path)

    elif path.endswith("xls"):

        def getBGColor(book, sheet, row, col):
            xfx = sheet.cell_xf_index(row, col)
            xf = book.xf_list[xfx]
            bgx = xf.background.pattern_colour_index
            pattern_colour = book.colour_map[bgx]

            # Actually, despite the name, the background colour is not the background colour.
            # background_colour_index = xf.background.background_colour_index
            # background_colour = book.colour_map[background_colour_index]

            return pattern_colour

        wb = xlrd.open_workbook(path, formatting_info=True)
        pattern_colour = getBGColor(wb, wb.sheet_by_index(0), cell[1][0] - 1, cell[1][1])

        out_wb = xlutils.copy.copy(wb)
        sheet = out_wb.get_sheet(0)

        if color:
            xlwt.add_palette_colour("font_colour", 0x21)
            out_wb.set_colour_RGB(0x21, *tuple(int(color[i:i+2], 16) for i in (0, 2, 4)))
            style = xlwt.easyxf('font: colour font_colour')
            sheet.write(cell[1][0] - 1, cell[1][1], value, style)
        else:
            sheet.write(cell[1][0] - 1, cell[1][1], value)

        """if pattern_colour:
            xlwt.add_palette_colour("cell_colour", 0x16)
            out_wb.set_colour_RGB(0x16, *pattern_colour)
            style = xlwt.easyxf('pattern: pattern solid, fore_colour cell_colour;'
                                'font: colour font_colour')
        else:
            style = xlwt.easyxf('font: colour font_colour')"""
        out_wb.save(path)






