from dataclasses import dataclass

from openpyxl import load_workbook
import numpy as np
from typing import Dict
from scipy.interpolate import interp1d, splrep, splev, make_interp_spline, BSpline, pchip_interpolate, griddata
import pyexcel as p
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
from general.general_functions import sigmoida

propertyPosition = {
    "laboratory_number": "A",
    "borehole": 'B',
    "depth": 'C',
    "soil_name":'D',
    "ige": 'ES',
    "rs": 'P',
    "r": 'Q',
    "rd": 'R',
    "n": 'S',
    "e": 'T',
    "W": 'U',
    "Sr": 'V',
    "Wl": 'W',
    "Wp": 'X',
    "Ip": 'Y',
    "Il": 'Z',
    "Ir": 'AE',
    "stratigraphic_index": 'AH',
    "ground_water_depth": 'AJ',
    "build_press": 'AK',
    "pit_depth": 'AL',
    "granulometric_10": 'E',
    "granulometric_5": 'F',
    "granulometric_2": 'G',
    "granulometric_1": 'H',
    "granulometric_05": 'I',
    "granulometric_025": 'J',
    "granulometric_01": 'K',
    "granulometric_005": 'L',
    "granulometric_001": 'M',
    "granulometric_0002": 'N',
    "granulometric_0000": 'O',
    "magnitude": "AQ",
    "intensity": "AM",
    "Cv": "CC",
    "Ca": "CF",
}

def generator_of_cell_with_lab_number(wb):
    """Функция генерирует последовательность строк с заполненными данными по лабномеру"""
    for i in range(7, len(wb['Лист1']['A']) + 5):
        if str(wb["Лист1"]['A' + str(i)].value) != "None":
            yield i


@dataclass
class BaseProperties:
    """Класс, хранящий свойсва грунтов, которые считываются без обработки"""
    laboratory_number: str = None
    borehole: str = None
    depth: float = None
    soil_name: str = None
    ige: str = None
    rs: float = None
    r: float = None
    rd: float = None
    n: float = None
    e: float = None
    W: float = None
    Sr: float = None
    Wl: float = None
    Wp: float = None
    Ip: float = None
    Il: float = None
    Ir: float = None
    stratigraphic_index: str = None
    ground_water_depth: float = None
    build_press: float = None
    pit_depth: float = None
    granulometric_10: float = None
    granulometric_5: float = None
    granulometric_2: float = None
    granulometric_1: float = None
    granulometric_05: float = None
    granulometric_025: float = None
    granulometric_01: float = None
    granulometric_005: float = None
    granulometric_001: float = None
    granulometric_0002: float = None
    granulometric_0000: float = None
    complete_flag: bool = False
    sample_number: int = None
    type_ground: int = None

    def defineBaseProperties(self, work_book, string, identification_column) -> None:
        """Считывание строки свойств"""
        for i, attr_name in enumerate(self.__dict__):
            if attr_name == "complete_flag":
                if identification_column:
                    cell = wb["Лист1"][identification_column + str(string)]
                    color_in_hex = cell.fill.start_color.index  # this gives you Hexadecimal value of the color
                    #color = tuple(int(color_in_hex[i:i + 2], 16) for i in (0, 2, 4))
                    if color_in_hex == "FF81D8D0":
                        self.complete_flag = True
                    else:
                        self.complete_flag = False
                else:
                    self.complete_flag = False

            elif attr_name == "sample_number":
                setattr(self, attr_name, string)
            elif attr_name == "laboratory_number":
                new_lab_number = str(wb["Лист1"]['IG' + str(string)].value)
                if new_lab_number != "None" and new_lab_number != "":
                    self.laboratory_number = new_lab_number
                else:
                    self.laboratory_number = str(wb["Лист1"]['A' + str(string)].value)
            elif attr_name == "type_ground":
                self.type_ground = PhysicalProperties.define_type_ground(self._granulometric_to_dict(), self.Ip,
                                                                         self.Ir)
            else:
                try:
                    setattr(self, attr_name,
                        BaseProperties.float_cell(work_book["Лист1"][propertyPosition[attr_name] + str(string)].value))
                except KeyError:
                    setattr(self, attr_name, None)

        self.proccessing_properties()

    def proccessing_properties(self) -> None:
        """Обработка незаполненых свойств"""
        self.e = np.random.uniform(0.55, 0.7) if not self.e else np.round(self.e, 2)
        if not self.Il:
            self.Il = np.random.uniform(0.4, 0.6) if not self.Ip else np.random.uniform(-0.1, 0.2)

    @staticmethod
    def define_type_ground(data_gran, Ip, Ir):
        """Функция определения типа грунта через грансостав"""
        none_to_zero = lambda x: 0 if not x else x
        gran_struct = ['10', '5', '2', '1', '05', '025', '01', '005', '001', '0002', '0000']  # гран состав
        accumulate_gran = [none_to_zero(data_gran[gran_struct[0]])]  # Накоплено процентное содержание
        for i in range(10):
            accumulate_gran.append(accumulate_gran[i] + none_to_zero(data_gran[gran_struct[i + 1]]))

        if none_to_zero(Ir) >= 50:  # содержание органического вещества Iom=hg10=Ir
            type_ground = 9  # Торф
        elif none_to_zero(Ip) < 1:  # число пластичности
            if accumulate_gran[2] > 25:
                type_ground = 1  # Песок гравелистый
            elif accumulate_gran[4] > 50:
                type_ground = 2  # Песок крупный
            elif accumulate_gran[5] > 50:
                type_ground = 3  # Песок средней крупности
            elif accumulate_gran[6] >= 75:
                type_ground = 4  # Песок мелкий
            else:
                type_ground = 5  # Песок пылеватый
        elif 1 <= Ip < 7:
            type_ground = 6  # Супесь
        elif 7 <= Ip < 17:
            type_ground = 7  # Суглинок
        else:  # data['Ip'] >= 17:
            type_ground = 8  # Глина
        if none_to_zero(Ir) >= 10:
            type_ground = 9

        return type_ground

    @staticmethod
    def float_cell(x):
        if x:
            try:
                return float(x)
            except ValueError:
                return x
        else:
            return None

@dataclass
class PhysicalProperties(BaseProperties):
    """Расширенный класс с дополнительными обработанными свойствами"""
    Cv: float = None
    Ca: float = None
    m: float = None

    def set_excel_file(self, work_book, string, identification_column=None) -> None:
        """Считывание строки свойств"""
        self.defineBaseProperties(work_book, string, identification_column)
        self.m = PhysicalProperties.define_m(self.e, self.Il)
        self.Cv = self.Cv if self.Cv else np.round(PhysicalProperties.define_Cv(
            PhysicalProperties.define_kf(self.type_ground, self.e)), 3)
        self.Ca = self.Ca if self.Ca else np.round(np.random.uniform(0.01, 0.03), 5)

        print(self.__dict__)

    def _granulometric_to_dict(self):
        granulometric_dict = {}
        for key in ['10', '5', '2', '1', '05', '025', '01', '005', '001', '0002', '0000']:
            granulometric_dict[key] = getattr(self, "granulometric_" + key)
        return granulometric_dict

    @staticmethod
    def define_m(e, Il):
        """Функция расчета параметра m - показатель степени для зависимости жесткости от уровня напряжений
         Входные параметры:
            :param e: пористость
            :param Il: число пластичности"""
        if (not Il or not e):
            return np.round(np.random.uniform(0.5, 0.65), 2)
        else:
            default_Il = [-0.25, 0, 0.25, 0.5, 0.75, 1]
            default_e = [0.5, 0.8, 1.2]
            default_m = [0.5, 0.6, 0.6, 0.7, 0.8, 0.8, 0.4, 0.5, 0.6, 0.7, 0.7, 0.9, 0.2, 0.3, 0.4, 0.6, 0.8, 1.0]

            if Il < default_Il[0]:
                Il = default_Il[0]
            if Il > default_Il[-1]:
                Il = default_Il[-1]
            if e < default_e[0]:
                e = default_e[0]
            if e > default_e[-1]:
                e = default_e[-1]

            default_e_Il = []
            for _e in default_e:
                default_e_Il.extend(list(map(lambda Il: [_e, Il], default_Il)))

            m = griddata(default_e_Il, default_m, (e, Il), method='cubic').item()

            try:
                return round(m, 2)
            except:
                round(np.random.uniform(0.5, 0.65), 2)

    @staticmethod
    def define_kf(type_ground: int, e) -> float:
        """ Определение коэффициента фильтрации по грансоставу
            :param type_ground: тип грунта
            :param e: коэффициент пористости
            :return: kf в метрах/сутки"""

        # Функция сигмоиды для kf
        kf_sigmoida = lambda e, e_min, e_max, k_min, k_max: sigmoida(e, amplitude=(k_max - k_min) / 2,
                                                                     x_indent=e_min + (e_max - e_min) / 2,
                                                                     y_indent=k_min + (k_max - k_min) / 2,
                                                                     shape=e_max - e_min)
        # Общие параметры сигмоиды
        e_borders = [0.3, 1.2]

        # Зависимость коэффициента фильтрации от грансостава
        dependence_kf_on_type_ground = {
            1: kf_sigmoida(e, *e_borders, 8.64, 86.4),
            2: kf_sigmoida(e, *e_borders, 8.64, 86.4),
            3: kf_sigmoida(e, *e_borders, 0.864, 86.4),
            4: kf_sigmoida(e, *e_borders, 8.64 * 10 ** (-2), 0.864),
            5: kf_sigmoida(e, *e_borders, 8.64 * 10 ** (-2), 0.864),
            6: kf_sigmoida(e, *e_borders, 8.64 * 10 ** (-4), 8.64 * 10 ** (-2)),
            7: kf_sigmoida(e, *e_borders, 8.64 * 10 ** (-5), 8.64 * 10 ** (-4)),
            8: kf_sigmoida(e, *e_borders, 0.0000001, 8.64 * 10 ** (-5))
        }

        return dependence_kf_on_type_ground[type_ground]

    @staticmethod
    def define_Cv(kf: float, m: float = 0.6) -> float:
        """ Определение коэффициента первичной консолидации Сv в см^2/мин
            :param kf: коэффициент фильтрации
            :param m: коэффициент относительной сжимаемости
            :return: Cv в см^2/мин"""

        # Переведем м/сут в см/мин
        kf *= 0.0694444

        # Переведем 1/МПа в 1 / (кгс / см2)
        m /= 10.197162

        # Удельный вес воды в кгс/см3
        gamma = 0.001

        Cv = kf / (m * gamma)

        if Cv > 1.2:
            return np.round(np.random.uniform(0.7, 1.2), 4)
        elif Cv <= 0.02:
            return np.round(np.random.uniform(0.01, 0.02), 4)
        return np.round(Cv, 4)


wb = load_workbook("C:/Users/Пользователь/Desktop/Тест/818-20 Атомфлот - мех.xlsx", data_only=True)

a = PhysicalProperties()
a.set_excel_file(wb, 7, "A")







