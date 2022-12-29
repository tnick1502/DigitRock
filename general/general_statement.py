import copy
from datetime import datetime
from typing import Tuple

from PyQt5.QtWidgets import QFileDialog, QHBoxLayout, QGroupBox, QDialog, \
    QComboBox, QWidget, QLineEdit, QPushButton, QVBoxLayout, QLabel, QMessageBox, QApplication, QCheckBox
from PyQt5.QtCore import Qt
import sys
import os

import numpy as np

from openpyxl import load_workbook

# from general.excel_functions import read_customer, form_xlsx_dictionary, table_data
from excel_statment.params import accreditation
from general.general_functions import create_json_file, read_json_file, unique_number, number_format
from general.initial_tables import Table
from general.report_general_statment import save_report
import xlrd
from openpyxl.utils import get_column_letter, column_index_from_string

from general.report_writer import ReportXlsxSaver
from singletons import statment


def convert_data(data):
    def zap(val, prec, none='-'):
        """ Возвращает значение `val` в виде строки с `prec` знаков после запятой
        используя запятую как разделитель дробной части
        """
        if isinstance(val, str):
            return val
        if val is None:
            return none
        fmt = "{:." + str(int(prec)) + "f}"
        return fmt.format(val).replace(".", ",")

    def val_to_list(val, prec) -> list:
        if val is None:
            return None
        else:
            try:
                val = [float(val)]
            except ValueError:
                v = val.split(";")
                val = []
                for value in v:
                    try:
                        a = float(value.replace(",", ".").strip(" "))
                        a = zap(a, prec)
                        val.append(a)
                    except:
                        pass

            return val

    data_new = []

    for i in range(len(data)):
        try:
            line = data[i]
            try:
                borehole = float(line[1])
                if borehole % 1 < 0.001:
                    line[1] = str(int(borehole))
                else:
                    line[1] = str(borehole)
            except ValueError:
                pass

            line[2] = line[2].replace(".", ",")# zap(line[2], 1, none='-')

            for i in range(3, len(line)):
                try:
                    line[i] = line[i].replace(".", ",")
                except:
                    pass

        except IndexError:
            pass

        try:
            if len(val_to_list(line[5], 1)) > 0:
                f = val_to_list(line[4], 1)
                E50 = val_to_list(line[5], 1)
                Ed = val_to_list(line[6], 1)
                Kd =val_to_list(line[7], 2)

                line = [line[0], line[1], line[2], line[3], f[0], E50[0], Ed[0], Kd[0]]
                data_new.append(line)

                for j in range(1, len(f)):
                    line = [line[0], line[1], line[2], line[3], f[j], E50[j], Ed[j], Kd[j]]
                    data_new.append(line)
            else:
                data_new.append(line)
        except:
            data_new.append(line)

    return data_new



    #x.insert(val, pos)

def convert_data2(data):
    def zap(val, prec, none='-'):
        """ Возвращает значение `val` в виде строки с `prec` знаков после запятой
        используя запятую как разделитель дробной части
        """
        if isinstance(val, str):
            return val
        if val is None:
            return none
        fmt = "{:." + str(int(prec)) + "f}"
        return fmt.format(val).replace(".", ",")

    def val_to_list(val, prec) -> list:
        if val is None:
            return None
        else:
            try:
                val = [float(val)]
            except ValueError:
                v = val.split(";")
                val = []
                for value in v:
                    try:
                        a = float(value.replace(",", ".").strip(" "))
                        a = zap(a, prec)
                        val.append(a)
                    except:
                        pass

            return val

    data_new = []

    for i in range(len(data)):
        try:
            line = data[i]
            try:
                borehole = float(line[1])
                if borehole % 1 < 0.001:
                    line[1] = str(int(borehole))
                else:
                    line[1] = str(borehole)
            except ValueError:
                pass

            line[2] = line[2].replace(".", ",")# zap(line[2], 1, none='-')

            for i in range(3, len(line)):
                try:
                    line[i] = line[i].replace(".", ",")
                except:
                    pass

        except IndexError:
            pass

        try:
            if len(val_to_list(line[4], 1)) > 0:
                f = val_to_list(line[4], 1)

                amp = zap(line[5], 1)
                rel = val_to_list(line[6], 2)
                alpha = zap(line[7], 3)
                betta = zap(line[8], 3)

                line = [line[0], line[1], line[2], line[3], f[0], amp, rel[0], alpha, betta]
                data_new.append(line)

                for j in range(1, len(f)):
                    line = [line[0], line[1], line[2], line[3], f[j], amp, rel[j], alpha, betta]
                    data_new.append(line)
            else:
                data_new.append(line)
        except:
            data_new.append(line)

    return data_new



    #x.insert(val, pos)



class StatementGenerator(QDialog):
    """
    Класс для представления пользовательского интерфейса импорта ведомости и
    вывода и экспорта обработанных данных в соответствии с заданными параметрами


    Атрибуты
    --------

    path : str
        путь к xls файлу ведомости
    customer : dict
        загруженные данные из ведомости по ключам
        ["customer", "object_name", "data", "accreditation"] о ["Заказчик", "Объект", "Дата", "Аккредитация"]
    statment_data : dict
        словарь с ключами по наименованиям колонок и соответствующими массивами колонок - numpy.ndarray


    Методы
    ------
    create_UI():
        устанавливает пользовательский интерфейс для импорта ведомости и вывода и экспорта
        считывает из ведомости данные в customer и statment_data

    _plot():
        выводит данные в таблицу на интерфесе
    _save_report():
        экспортирует данные в pdf-файл
    _structure_assretion_tests(table, structure):
        тесты для проверки корректности введенных пользователем параметров и возможности
        отображения и/или экспорта ведомости. Возвращает true, если все тесты успешны

    """

    def __init__(self, parent, path=None, statment_data=None, statement_structure_key=None,
                 test_mode_and_shipment: Tuple[str, str] = (None, None)):
        super().__init__(parent)

        self.sort = False

        self.setGeometry(100, 50, 1000, 950)

        self.path = path
        self.customer = None

        self.statment_data = statment_data

        self._statement_structure_key = statement_structure_key if statement_structure_key else "triaxial_cyclic"

        self.statment_test_mode, self.shipment = test_mode_and_shipment

        self.create_UI()

        if path:
            self.open_excel(path)

        if statement_structure_key:
            self._plot()


    def create_UI(self):
        self.layout = QVBoxLayout()
        self.layout.setSpacing(5)
        self.layout.setContentsMargins(5, 5, 5, 5)

        self.open_box = QGroupBox("Текущая ведомость")
        self.open_box_layout = QHBoxLayout()
        self.button_open = QPushButton("Открыть файл ведомости")
        self.button_open.clicked.connect(self.open_excel)
        self.open_box_layout.addWidget(self.button_open)
        self.text_file_path = QLineEdit()
        self.text_file_path.setDisabled(True)
        self.open_box.setFixedHeight(80)
        self.open_box_layout.addWidget(self.text_file_path)
        self.open_box.setLayout(self.open_box_layout)
        self.layout.addWidget(self.open_box)

        self.customer_table = Table(headers=["Заказчик", "Объект", "Дата", "Аккредитация"])
        self.customer_table.set_data([["Заказчик", "Объект", "Дата", "Аккредитация"], ["", "", "", ""]], "Stretch")
        self.customer_table.setFixedHeight(80)
        self.layout.addWidget(self.customer_table)

        self.StatementStructure = StatementStructure(statement_structure_key=self._statement_structure_key)
        self.layout.addWidget(self.StatementStructure)

        self.statment_table = Table(moove=True)
        self.layout.addWidget(self.statment_table)

        self.StatementStructure.plot_structure_button.clicked.connect(self._plot)

        self.StatementStructure.save_button.clicked.connect(self._save_report)
        self.StatementStructure.save_button_xls.clicked.connect(self._save_report_xls)


        self.StatementStructure.sort_btn.clicked.connect(self._on_sort)

        self.setLayout(self.layout)

    def open_excel(self, path=None):
        if path:
            self.path = path
        else:
            self.path = QFileDialog.getOpenFileName(self, 'Open file', '/home')[0]

        try:
            #wb = load_workbook(self.path, data_only=True)
            marker, self.customer = self.read_customer(self.path)
            self.customer_table.set_data([["Заказчик", "Объект", "Дата", "Аккредитация"],
                                          [self.customer[i] for i in
                                           ["customer", "object_name", "data", "accreditation"]]], "Stretch")
            self.text_file_path.setText(self.path)
            self.statment_data = self.form_excel_dictionary(self.path, last_key='IV')
            self.accreditation = self.customer["accreditation"]
            self.accreditation_key = "новая"

            if self.accreditation in ["OAO", "ОАО"]:
                self.accreditation = "АО"
            elif self.accreditation == "OOO":
                self.accreditation = "ООО"

            if self.accreditation in ["AO", "АО"]:
                self.accreditation_key = "новая"
            elif self.accreditation == "ООО" or self.accreditation == "OOO":
                self.accreditation_key = "2"

        except FileNotFoundError as error:
            print(error)

    def _plot(self):
        # print(self.StatementStructure.get_structure())
        # print(table_data(self.statment_data, self.StatementStructure.get_structure()))
        if self.statment_data:
            if self._structure_assretion_tests(self.statment_data, self.StatementStructure.get_structure()):
                titles, data, scales = self.table_data(self.statment_data, self.StatementStructure.get_structure())
                for i in range(len(data)):
                    for j in range(len(data[i])):
                        if data[i][j] == 'None':
                            data[i][j] = ' '

                if self.sort:
                    data = self.sort_data_by_skv_depth(data)

                self.statment_table.set_data([titles] + data, "Stretch")

            else:
                pass

    def _structure_assretion_tests(self, table, structure):
        '''
        функция проверки корректности структуры для имеющейся таблицы данных
        возвращает True если все тесты успешны
        возвращает False если нет
        '''
        try:

            # Блок теста триггеров:

            # Корректировка триггеров как в table_data
            if (structure["trigger"] is None) or (structure["trigger"] == []):
                structure["trigger"] = [None]
            while len(structure["trigger"]) > 1 and structure["trigger"].count(None) > 0:
                structure["trigger"].remove(
                    None)  # удаляем None так, чтобы остался массив из одного None на случай массива [None, A]
            # для каждого тригера вызываем проверку его налачия в данных таблицы
            if structure["trigger"].count(None) == 0:
                for i in range(len(structure["trigger"])):
                    assert (structure["trigger"][i] in table.keys()), 'Триггер ' + str(
                        structure["trigger"][i]) + ' отсутствует'

            #
            for i in range(len(structure["columns"])):
                # for j in range(len(structure["columns"][str(i)]['cell'])):
                assert (structure["columns"][str(i)]['cell'] in table.keys()), 'Ячейка ' + str(
                    structure["columns"][str(i)]['cell']) + ' отсутствует'

            return True

        except AssertionError as error:
            QMessageBox.critical(self, "Ошибка", str(error))
            return False

    def _save_report_xls(self):
        if self.statment_data:
            if self._structure_assretion_tests(self.statment_data, self.StatementStructure.get_structure()):
                # try:
                    # file = QFileDialog.getOpenFileName(self, 'Open file')[0]
                save_file_pass = QFileDialog.getExistingDirectory(self, "Select Directory")

                if self.statment_test_mode and self.shipment:
                    save_file_name = f'Ведомость {self.statment_test_mode} {self.shipment}.pdf'
                else:
                    save_file_name = 'Общая ведомость.pdf'
                # считывание параметра "Заголовок"

                statement_title = self.StatementStructure.get_structure().get("statement_title", '')

                # self.StatementStructure._additional_parameters = \
                #    StatementStructure.read_ad_params(self.StatementStructure.additional_parameters.text())
                titles, data, scales = self.table_data(self.statment_data, self.StatementStructure.get_structure())

                try:
                    if statment.general_parameters.test_mode == "Виброползучесть":
                       data = convert_data(data)
                    elif statment.general_parameters.test_mode == "Демпфирование по Релею":
                        data = convert_data2(data)
                except:
                    pass

                for i in range(len(data)):
                    for j in range(len(data[i])):
                        if data[i][j] == 'None':
                            data[i][j] = ' '
                if self.sort:
                    data = self.sort_data_by_skv_depth(data)
                # ["customer", "object_name", "data", "accreditation"]
                # ["Заказчик", "Объект", "Дата", "Аккредитация"]
                # Дата
                data_report = self.customer["data"]
                customer_data_info = ['Заказчик:', 'Объект:']
                # Сами данные (подробнее см. Report.py)
                customer_data = [self.customer[i] + "                  " for i in ["customer", "object_name"]]

                statement_title += f" №{self.customer['object_number']}СВД"

                try:
                    if save_file_pass:
                        accred1 = {'accreditation': self.accreditation,
                                   'accreditation_key': self.accreditation_key}
                        if accred1 is None:
                            accred1 = {'acrreditation': 'AO', 'acrreditation_key': 'новая'}
                        if accred1:
                            accred = [accreditation[accred1['accreditation']][accred1['accreditation_key']][0],
                                      accreditation[accred1['accreditation']][accred1['accreditation_key']][1]]
                        else:
                            accred = [
                                'АТТЕСТАТ АККРЕДИТАЦИИ №RU.MCC.АЛ.988 Срок действия с 09 января 2020г.',
                                'РЕЕСТР ГЕОНАДЗОРА г. МОСКВЫ №27 (РЕЙТИНГ №4)']

                        is_template = False
                        template_filename = None
                        if self.StatementStructure.combo_box.currentText() == 'vibriation_creep':
                            template_filename = 'xls_statment_VIBRO_template.xlsx'
                            num_page_rows = 55
                            is_template = True
                        elif self.StatementStructure.combo_box.currentText() == 'damping':
                            template_filename = 'xls_statment_DEMPH_template.xlsx'
                            num_page_rows = 54
                            is_template = True

                        elif self.StatementStructure.combo_box.currentText() == 'rayleigh_damping':
                            template_filename = 'xls_statment_RELEY_template.xlsx'
                            num_page_rows = 54
                            is_template = True

                        elif self.StatementStructure.combo_box.currentText() == 'Resonance column':
                            template_filename = 'xls_statment_RESONANT_template.xlsx'
                            num_page_rows = 54
                            is_template = True

                        elif self.StatementStructure.combo_box.currentText() == 'Seismic liquefaction':
                            template_filename = 'xls_statment_SEISMO_template.xlsx'
                            num_page_rows = 53
                            is_template = True

                        elif self.StatementStructure.combo_box.currentText() == 'Storm liquefaction':
                            template_filename = 'xls_statment_STORM_template.xlsx'
                            num_page_rows = 53
                            is_template = True

                        else:
                            is_template = False

                        # Если данные в полях не отличают от шаблона, то можно воспользоваться шаблоном
                        if not self.StatementStructure.is_template_changed() and is_template and template_filename:

                            # Здесь необходимо для других шаблонов прописать соответсвующие им файлы


                            writer = ReportXlsxSaver(template_filename=template_filename,
                                                     num_page_rows=num_page_rows)

                            formatted_tests_data, additional = writer.form_tests_data_list_mode(titles, data)

                            writer.set_data_row_mode(customer=customer_data[0],
                                                     obj_name=customer_data[1],
                                                     test_title=statement_title,
                                                     date=data_report,
                                                     tests_data=formatted_tests_data,
                                                     accreditation=f"{accred[0]}\n{accred[1]}",
                                                     additional_data=additional,
                                                     doc_num=unique_number(length=7, postfix="-СВД"))
                            writer.save(f"{save_file_pass}/{save_file_name.replace('.pdf','.xlsx')}")
                        else:
                            writer = ReportXlsxSaver()

                            formatted_tests_data, additional = writer.form_tests_data_dict_mode(titles, data)

                            writer.set_data_dict_mode(customer=customer_data[0],
                                                      obj_name=customer_data[1],
                                                      test_title=statement_title,
                                                      date=data_report,
                                                      tests_data=formatted_tests_data,
                                                      accreditation=f"{accred[0]}\n{accred[1]}",
                                                      additional_data=additional,
                                                      doc_num=unique_number(length=7, postfix="-СВД"))
                            writer.save(f"{save_file_pass}/{save_file_name.replace('.pdf', '.xlsx')}")
                        # save_report(titles, data, scales, data_report, customer_data_info, customer_data,
                        #             statement_title, save_file_pass, unique_number(length=7, postfix="-СВД"),
                        #             save_file_name, accred1={'accreditation': self.accreditation,
                        #                                     'accreditation_key': self.accreditation_key})
                        QMessageBox.about(self, "Сообщение", "Успешно сохранено")
                except PermissionError:
                    QMessageBox.critical(self, "Ошибка", "Закройте файл для записи", QMessageBox.Ok)
                # except (ValueError, IndexError, ZeroDivisionError) as error:
                #     QMessageBox.critical(self, "Ошибка", str(error), QMessageBox.Ok)
                #     pass
            else:
                pass
            pass

    def _save_report(self):
        if self.statment_data:
            if self._structure_assretion_tests(self.statment_data, self.StatementStructure.get_structure()):
                try:
                    # file = QFileDialog.getOpenFileName(self, 'Open file')[0]
                    save_file_pass = QFileDialog.getExistingDirectory(self, "Select Directory")

                    if self.statment_test_mode and self.shipment:
                        save_file_name = f'Ведомость {self.statment_test_mode} {self.shipment}.pdf'
                    else:
                        save_file_name = 'Общая ведомость.pdf'
                    # считывание параметра "Заголовок"

                    statement_title = self.StatementStructure.get_structure().get("statement_title", '')

                    # self.StatementStructure._additional_parameters = \
                    #    StatementStructure.read_ad_params(self.StatementStructure.additional_parameters.text())
                    titles, data, scales = self.table_data(self.statment_data, self.StatementStructure.get_structure())

                    try:
                        if statment.general_parameters.test_mode == "Виброползучесть":
                           data = convert_data(data)
                        elif statment.general_parameters.test_mode == "Демпфирование по Релею":
                            data = convert_data2(data)
                    except:
                        pass

                    for i in range(len(data)):
                        for j in range(len(data[i])):
                            if data[i][j] == 'None':
                                data[i][j] = ' '
                    if self.sort:
                        data = self.sort_data_by_skv_depth(data)
                    # ["customer", "object_name", "data", "accreditation"]
                    # ["Заказчик", "Объект", "Дата", "Аккредитация"]
                    # Дата
                    data_report = self.customer["data"]
                    customer_data_info = ['Заказчик:', 'Объект:']
                    # Сами данные (подробнее см. Report.py)
                    customer_data = [self.customer[i] + "                  " for i in ["customer", "object_name"]]

                    statement_title += f" №{self.customer['object_number']}СВД"

                    try:
                        if save_file_pass:
                            save_report(titles, data, scales, data_report, customer_data_info, customer_data,
                                        statement_title, save_file_pass, unique_number(length=7, postfix="-СВД"),
                                        save_file_name, accred1={'accreditation': self.accreditation,
                                                                'accreditation_key': self.accreditation_key})
                            QMessageBox.about(self, "Сообщение", "Успешно сохранено")
                    except PermissionError:
                        QMessageBox.critical(self, "Ошибка", "Закройте файл для записи", QMessageBox.Ok)
                    except:
                        pass
                except (ValueError, IndexError, ZeroDivisionError) as error:
                    QMessageBox.critical(self, "Ошибка", str(error), QMessageBox.Ok)
                    pass
            else:
                pass
            pass

    def generator_of_cell_with_lab_number_xls(self, sheet):
        """Функция генерирует последовательность строк с заполненными данными по лабномеру"""
        for i in range(6, sheet.nrows):
            if str(sheet.cell(i, 0).value).replace(' ', '') not in ["None", ""]:
                yield i

    def generator_of_cell_with_lab_number_xlsx(self, wb):
        """Функция генерирует последовательность строк с заполненными данными по лабномеру"""
        for i in range(7, len(wb['Лист1']['A']) + 5):
            if str(wb["Лист1"]['A' + str(i)].value) != "None":
                yield i

    def get_column_index(self, colname: str):
        for i in range(1000):
            if colname == xlrd.colname(i):
                return i
        return None

    def get_column_letters(self, last_letter='IV'):
        """
        Функция формирует список наименований колонок из exel (ключей)
        :param last_letter: str, название колонки по которую (включительно) формировать словарь
        :return: list, список наименований колонок (ключи)
        """
        import_columns = []
        try:
            last_letter_index = column_index_from_string(last_letter)
        except:
            last_letter_index = column_index_from_string('IV')
        for i in range(1, last_letter_index + 1):
            import_columns.append(get_column_letter(i))
        return import_columns[:last_letter_index + 1]  # обрезание идет До индекса, так что включаем еще значение

    def form_xls_dictionary(self, sheet, last_key='IV'):
        """

        """

        # наименования колонок большими буквами
        last_key = last_key.upper()

        # sheet = sheet.sheet_by_index(0)

        # формируем список ключей
        import_columns = self.get_column_letters(last_key)

        # объявляем словарь
        xls_dictionary = {str(import_columns[0]): []}
        # вносим в него ключи
        for col in import_columns:
            xls_dictionary[str(col)] = []

        ig_index = self.get_column_index('IG')
        a_index = self.get_column_index('A')

        for col in range(len(import_columns)):
            for row in self.generator_of_cell_with_lab_number_xls(sheet):
                # выполняем проверку, что ячейка не пустая
                if str(sheet.cell(row, col).value).replace(' ', '') not in ["None", ""]:
                    if xlrd.colname(col) == "A":
                        if str(sheet.cell(row, ig_index).value) not in ["None", ""]:
                            xls_dictionary[xlrd.colname(col)].append(str(sheet.cell(row, ig_index).value))
                        else:
                            xls_dictionary[xlrd.colname(col)].append(str(sheet.cell(row, a_index).value))
                    else:
                        xls_dictionary[xlrd.colname(col)].append(str(sheet.cell(row, col).value))
                else:
                    xls_dictionary[xlrd.colname(col)].append("None")

        for key in xls_dictionary:
            xls_dictionary[str(key)] = np.array(xls_dictionary[str(key)])

        return xls_dictionary

    def form_xlsx_dictionary(self, wb, last_key):
        """
        Функция считывает всю ведомость и записывает значения в словарь, где
        ключи - обозначения колонок в exel таблице ('A', 'B', ...)
        значения - колонок из ведомости в виде массивов
        :param wb: workbook, результат импорта ведомости
        :param last_key: str, название колонки (ключа) по которую (включительно) формировать словарь
        :return: dict, словарь с ключами по наименованиям колонок и соответствующими массивами колонок - numpy.ndarray
        """

        # наименования колонок большими буквами
        last_key = last_key.upper()

        # формируем список ключей
        import_columns = self.get_column_letters(last_key)

        # объявляем словарь
        xlsx_dictionary = {str(import_columns[0]): []}
        # вносим в него ключи
        for col in import_columns:
            xlsx_dictionary[str(col)] = []

        for key in xlsx_dictionary:
            for i in self.generator_of_cell_with_lab_number_xlsx(wb):
                # выполняем проверку, что ячейка не пустая
                if str(wb["Лист1"][str(key) + str(i)].value) != "None":
                    if str(key) == "A":
                        if str(wb["Лист1"]["IG" + str(i)].value) not in ["None", ""]:
                            xlsx_dictionary[str(key)].append(wb["Лист1"]["IG" + str(i)].value)
                        else:
                            xlsx_dictionary[str(key)].append(wb["Лист1"]["A" + str(i)].value)
                    else:
                        xlsx_dictionary[str(key)].append(wb["Лист1"][str(key) + str(i)].value)
                else:
                    xlsx_dictionary[str(key)].append("None")
        # переводим значения в массивы numpy.ndarray
        for key in xlsx_dictionary:
            xlsx_dictionary[str(key)] = np.array(xlsx_dictionary[str(key)])

        return xlsx_dictionary

    def form_excel_dictionary(self, path, last_key='IV'):

        # наименования колонок большими буквами
        last_key = last_key.upper()
        if path[-1] == "x":
            #print("xlsx")
            sheet = load_workbook(path, data_only=True)
            return self.form_xlsx_dictionary(sheet, last_key)

        else:
            #print("xls")
            sheet = xlrd.open_workbook(path, formatting_info=True)
            sheet = sheet.sheet_by_index(0)
            return self.form_xls_dictionary(sheet, last_key)

    def read_customer(self, path):
        """Чтение данных заказчика, даты
            Передается документ excel, возвращает маркер False и данные, либо маркер True и имя ошибки"""

        if path.endswith("xlsx"):
            wb = load_workbook(path, data_only=True)
            data = {"customer": str(wb["Лист1"]["A1"].value),
                    "object_name": str(wb["Лист1"]["A2"].value),
                    "data": wb["Лист1"]["Q1"].value,
                    "start_date": wb["Лист1"]["U1"].value,
                    "accreditation": str(wb["Лист1"]["I2"].value),
                    "object_number": str(wb["Лист1"]["AI1"].value)}
        else:
            wb = xlrd.open_workbook(path, formatting_info=True)
            wb = wb.sheet_by_index(0)
            data = {"customer": self.str_float(wb.cell(0, 0).value),
                    "object_name": self.str_float(wb.cell(1, 0).value),
                    "data": self.date_datetime(wb.cell(0, 16).value),
                    "start_date": self.date_datetime(wb.cell(0, 20).value),
                    "accreditation": self.str_float(wb.cell(1, 8).value),
                    "object_number": self.str_float(wb.cell(0, 34).value)}

        for i in data:
            if data[i] == "None":
                return True, i

        if not isinstance(data["data"], datetime):
            return True, "Дата окончания опытов"

        if not isinstance(data["start_date"], datetime):
            return True, "Дата начала опытов"

        return False, data

    def table_data(self, table, structure):
        """Функция возвращает матрицу для построения таблицы. Первая втрока - имя, остальные - столбцы значений
        Входные параметры: table - матрица, считанная с excel,
                           structure - словарь, описывающий структуру таблицы
                           structure = {"trigger": ['BI', 'AS'],
                           "columns": {"0": {"title": "Скважина", "cell": "B"},
                                      "1": {"title": "Лаб.номер", "cell": "A"},
                                      "2": {"title": "Глубина", "cell": "C"}}}"""

        data = [[]]

        titles = [structure["columns"][str(i)]["title"] for i in range(len(structure["columns"]))]

        scale = [structure["columns"][str(i)]["scale_factor"] for i in range(len(structure["columns"]))]

        parameter_decimal = [structure["columns"][str(i)]["number_of_decimal_places"] for i in
                             range(len(structure["columns"]))]

        # for i in range(len(structure["columns"])): # идем по строкам columns
        #     data[0].append(structure["columns"][str(i)]["title"])  # в список списков в первый список записываем все title из columns str(i)-дает обращение к нужному ключу по порядку

        if (structure["trigger"] is None) or (structure["trigger"] == []):
            structure["trigger"] = [None]
        while len(structure["trigger"]) > 1 and structure["trigger"].count(None) > 0:
            structure["trigger"].remove(
                None)  # удаляем None так, чтобы остался массив из одного None на случай массива [None, A]

        if structure["trigger"].count(None) == 0:
            k = 0
            for i in range(len(table[structure["trigger"][0]])):  # Идем по столбцу тригера
                flag = 1
                for tr in range(len(structure["trigger"])):  # внутренний цикл для того чтобы идти по всем тригерам
                    if table[structure["trigger"][tr]][
                        i] == 'None':  # если в ячейке из столбца тригера пусто меняем флаг на 0
                        flag = 0
                if flag:  # для непустых ячеек тригера дбавляем в список списков пустые списки
                    for j in range(len(structure["columns"])):  # по длине массива с названими
                        data[k].append(table[structure["columns"][str(j)]["cell"]][
                                           i])  # записываем в каждый массив значения из каждой строки в нужных столбцах
                    data.append([])
                    k += 1  # считаем количество строк которые были записаны
            data.pop(-1)
        else:  # если тригера нет
            for i in range(len(table[structure["columns"]["0"]["cell"]])):  # идем по столбцу первому из columns
                for j in range(len(structure["columns"])):
                    data[i].append(table[structure["columns"][str(j)]["cell"]][
                                       i])  # записываем в каждый массив значения из каждой строки в нужных столбцах
                data.append([])  # добавляем нужное число пустых массивов
            data.pop(-1)
        # print('до округ',data)
        # Форматирование данных под число знаком после запятой
        data = self.number_of_decimal_places(data, parameter_decimal)

        # Перевод scale к виду scales = [3 * cm, 5 * cm, 6 * cm]
        for i in range(len(scale)):
            try:
                scale[i] = 28.346456692913385 * float(
                    scale[i])  # значение cm в from reportlab.lib.units import inch, cm, mm
            except ValueError:
                pass
        s = [i.strip(" ") for i in structure["additional_parameter"].split(";")]
        #print('f11', s)
        for i in range(len(s)):
            data.append([s[i]])
        # data.append([structure["additional_parameter"]])

        #print('titles', titles)
        #print('data', data)
        #print('scale', scale)
        return titles, data, scale

    def number_of_decimal_places(self, matrix, parameter_decimal):
        for j, count in enumerate(parameter_decimal):
            if str(count) != 'None' and str(count) != '*' and str(count) != '':
                for i in range(len(matrix)):
                    try:
                        matrix[i][j] = number_format(self.float_float(matrix[i][j]),
                                                          characters_number=int(count), split=',')

                    except:
                        pass

        # !!!
        for i in range(len(matrix)):
            for j in range(len(matrix[0])):
                try:
                    matrix[i][j] = str(matrix[i][j])
                except:
                    pass

        return matrix

    def float_float(self, a):
        try:
            a = float(a)
            return a
        except ValueError:
            try:
                a = float(a.replace(",", '.'))
                return a
            except ValueError:
                return a

    def date_datetime(self, val):
        value = str(val)
        if not value.replace(' ', ''):
            return 0

        try:
            value = int(val)
        except ValueError:
            value = 0

        if value:
            return xlrd.xldate_as_datetime(value, 0)

        return 0

    def str_float(self, val):
        value = str(val)
        if value == " ":
            return "None"
        return value

    def _on_sort(self, checked: bool):
        if not checked:
            self.sort = False
        else:
            self.sort = True

        self._plot()

    def sort_data_by_skv_depth(self, data):
        result = copy.deepcopy(data[:-1])

        is_str_item_1 = False
        for item in result:
            if not is_str_item_1 and type(self.float_float(item[1])) == str:
                is_str_item_1 = True

        result = sorted(result, key=lambda x: [x[1] if is_str_item_1 else self.float_float(x[1]),
                                               self.float_float(x[2])] if len(x) > 2 else x[0])

        return [*result, data[-1]]

class StatementStructure(QWidget):
    """
    Класс для представления пользовательского интерфейса и механизмов создания и хранения шаблонов
    для параметров общей ведомости и структуры данных


    Атрибуты
    --------

    params : dict
        общий перечень перечень параметров ведомости
    _statement_structures_path : str
        путь к .json файлу с шаблономи ведомостей
    _statement_structures : dict
        общий перечень шаблонов параметров ведомости, ключем к конкретной структуре с шаблоном является имя шаблона
    _statement_structure : dict
        шаблон параметров ведомости в виде структуры вида:
            {"trigger": ["A"], "columns": {"0": {"title": "Скважина", "cell": "B", "number_of_decimal_places": None, "scale_factor": "*"},
                        "1": {"title": "Лаб.номер", "cell": "A", "number_of_decimal_places": None, "scale_factor": "*"},
                        "2": {"title": "Глубина", "cell": "C", "number_of_decimal_places": None, "scale_factor": "*"}}}


    Методы
    ------
    create_UI():
        устанавливает пользовательский интерфейс для параметров общей ведомости и вывода и сохранения результатов
    get_structure():
        считывает текущие параметры ведомости с интерфеса в структуру _statement_structure и возвращает её

    _open_statement_structures(path = None):
        загружает файл с шаблонами параметров в _statement_structures, загружает перечень шаблонов в список
    _combo_changed():
        устанавливает в _statement_structure текущий набор параметров по выбранному шаблону ведомости в списке шаблонов
    _set_combo_structure(key):
        устанавливает текущий шаблон по имени из списка
    _set_structure():
        заполняет поля параметров на интерфейсе согласно текущему набору в _statement_structure
    _save_structure():
        добавляет заполненные пользователем параметры к списку шаблонов по названию, определенному польлзователем
    _get_structure():
        загружает текущие параметры с интерфейса в _statement_structure

    """

    def __init__(self, path=None, statement_structure_key=None):
        super().__init__()

        self.params = {"parameter_title": "Заголовок",
                       "parameter_trigger": "Триггеры",
                       "parameter_cells": "Выбранные ячейки",
                       "parameter_column_titles": "Имена в ведомости",
                       "parameter_decimal": "Число знаков после запятой",
                       "scale_factor": "Размер столбцов",
                       "additional_parameters": "Дополнительные параметры испытаний"}

        self._statement_structures_path = os.path.join(os.getcwd() + "/project_data/", "structures.json")

        self._statement_structures = None
        self._statement_structure = None

        self.create_UI()
        self.setFixedHeight(38 * len(self.params) + 70)  # задаем высоту в завивисиости от числа параметров

        self._open_statement_structures(self._statement_structures_path)  # вызываем функцию от пути которая считывает структуру из файла json
        if statement_structure_key:  # только если в переменную передали ключ
            self._set_combo_structure(statement_structure_key)

    def create_UI(self):
        self.layout = QVBoxLayout()
        self.layout.setSpacing(5)
        self.layout.setContentsMargins(5, 5, 5, 5)

        self.parameter_box = QGroupBox("Параметры общей ведомости")
        self.parameter_box_layout = QVBoxLayout()

        for param in self.params.keys():
            setattr(self, "line_{}".format(param), QHBoxLayout())  # Создаем элемент QHBoxLayout()
            label = QLabel(self.params[param])  # Создааем подпись
            label.setFixedWidth(150)  # Фиксируем размер подписи
            getattr(self, "line_{}".format(param)).addWidget(label)  # Размещаем подпись на ранее созданном layout

            setattr(self, "{}".format(param), QLineEdit())  # Создаем элемент QLineEdit()
            getattr(self, "line_{}".format(param)).addWidget(getattr(self, "{}".format(param)))  # Размещаем
            # QLineEdit()
            # в layout
            self.parameter_box_layout.addLayout(getattr(self, "line_{}".format(param)))

        self.parameter_box.setLayout(self.parameter_box_layout)

        self.end_line = QHBoxLayout()
        self.dafault_parameter_box = QGroupBox("Шаблоны ведомостей")
        self.dafault_parameter_box.setFixedWidth(800)
        self.dafault_parameter_box.setFixedHeight(70)
        self.dafault_parameter_box_layout = QHBoxLayout()
        self.combo_box = QComboBox()
        self.combo_box.activated.connect(self._combo_changed)
        self.dafault_parameter_box_layout.addWidget(self.combo_box)
        self.combo_box.setFixedWidth(180)
        self.new_statement_name = QLineEdit()
        #self.new_statement_name.setFixedWidth(120)
        self.dafault_parameter_box_layout.addWidget(self.new_statement_name)
        self.save_new_structure_button = QPushButton("Сохранить шаблон")
        self.save_new_structure_button.setFixedWidth(110)
        self.dafault_parameter_box_layout.addWidget(self.save_new_structure_button)
        self.save_new_structure_button.clicked.connect(self._save_structure)
        self.open_new_structure_button = QPushButton("Открыть шаблон")
        self.open_new_structure_button.setFixedWidth(110)
        self.dafault_parameter_box_layout.addWidget(self.open_new_structure_button)
        self.dell_structure_button = QPushButton("Удалить шаблон")
        self.dell_structure_button.setFixedWidth(110)
        self.dafault_parameter_box_layout.addWidget(self.dell_structure_button)
        """!!!"""
        self.dell_structure_button.clicked.connect(self._dell_structure)

        self.dafault_parameter_box.setLayout(self.dafault_parameter_box_layout)
        self.end_line.addWidget(self.dafault_parameter_box)

        self.plot_structure_button = QPushButton("Построить по шаблону")
        self.plot_structure_button.setFixedWidth(140)
        self.plot_structure_button.setFixedHeight(70)
        self.end_line.addWidget(self.plot_structure_button)

        self.save_btns = QGroupBox()
        self.save_btns_layout = QVBoxLayout()
        self.save_btns.setLayout(self.save_btns_layout)

        self.save_button = QPushButton("Сохранить pdf")
        self.save_button.setFixedWidth(140)
        self.save_button.setFixedHeight(25)

        self.save_button_xls = QPushButton("Сохранить xls")
        self.save_button_xls.setFixedWidth(140)
        self.save_button_xls.setFixedHeight(25)

        self.save_btns_layout.addWidget(self.save_button)
        self.save_btns_layout.addWidget(self.save_button_xls)

        self.end_line.addWidget(self.save_btns)

        self.sort_btn = QCheckBox("Сортировка по скв/глуб")
        self.sort_btn.setFixedHeight(30)
        self.sort_btn.setFixedWidth(150)
        self.end_line.addWidget(self.sort_btn)

        self.end_line.addStretch(-1)
        self.parameter_box_layout.addLayout(self.end_line)
        self.layout.addWidget(self.parameter_box)
        self.setLayout(self.layout)

    def _open_statement_structures(self, path=None):
        """Чтение файла структур"""
        if path:
            file = path
        else:
            file = QFileDialog.getOpenFileName(self, 'Open file', '/home')[0]

        self._statement_structures_path = file

        try:
            self.combo_box.clear()  # необходимо для очистки предыдущего импорта если он был
            self._statement_structures = read_json_file(self._statement_structures_path)
            self.combo_box.addItems(self._statement_structures.keys())
        except:
            pass

    def _combo_changed(self):
        """Смена значений в combo_change"""
        if self._statement_structures:
            self._statement_structure = self._statement_structures[self.combo_box.currentText()]
        self._set_structure()
        """!!!"""
        self.new_statement_name.setText(self.combo_box.currentText())

    def _set_combo_structure(self, key):
        """Поставить значение по ключу в combo_box"""
        index = self.combo_box.findText(key, Qt.MatchFixedString)
        if index >= 0:
            self.combo_box.setCurrentIndex(index)
        if index == -1:
            self._statement_structure = None
        self._combo_changed()

    def _set_structure(self):
        """Заполнение формы и заголовка по структуре таблицы"""
        if self._statement_structure:
            statement_title, triggers, cells, titles, decimal, scale_factor, additional_parameters = StatementStructure.form_output_from_structure(
                self._statement_structure)
            self.parameter_title.setText(statement_title)
            self.parameter_trigger.setText(triggers)
            self.parameter_cells.setText(cells)
            self.parameter_column_titles.setText(titles)
            self.parameter_decimal.setText(decimal)
            self.scale_factor.setText(scale_factor)
            self.additional_parameters.setText(additional_parameters)
        else:
            self.parameter_title.setText("")
            self.parameter_trigger.setText("")
            self.parameter_cells.setText("")
            self.parameter_column_titles.setText("")
            self.parameter_decimal.setText("")
            self.scale_factor.setText("")
            self.additional_parameters.setText("")

        # self.additional_parameters.setText('; '.join(self._additional_parameters))

    def _save_structure(self):
        """Функция сохранения новой структуре в json файле"""
        text = self.new_statement_name.text()

        if text:
            try:
                # self._additional_parameters=StatementStructure.read_ad_params(self.additional_parameters.text())
                self._get_structure()

                self._statement_structures[text] = self._statement_structure
                create_json_file(self._statement_structures_path, self._statement_structures)
                self._open_statement_structures(self._statement_structures_path)
                self._set_combo_structure(text)
            except:
                QMessageBox.critical(self, "Ошибка", "Ошибка добавления")
        else:
            QMessageBox.critical(self, "Ошибка", "Введите имя шаблона")

    """!!!"""
    def _dell_structure(self):
        """Функция удаления структуры"""

        text = self.new_statement_name.text()

        if text:
            try:
                self._statement_structures.pop(text)
                create_json_file(self._statement_structures_path, self._statement_structures)
                self._open_statement_structures(self._statement_structures_path)
                self.combo_box.setCurrentIndex(0)
                self._set_combo_structure(self.combo_box.currentText())
            except KeyError:
                QMessageBox.critical(self, "Ошибка", "Неверное имя шаблона")

    def _get_structure(self):
        statement_title = self.parameter_title.text()
        triggers = StatementStructure.read_line(self.parameter_trigger.text())
        cells = StatementStructure.read_line(self.parameter_cells.text())
        titles = StatementStructure.read_titles(self.parameter_column_titles.text())
        decimal = StatementStructure.read_line(self.parameter_decimal.text())
        scale_factor = StatementStructure.read_scale(self.scale_factor.text())
        additional_parameters = self.additional_parameters.text()
        self._statement_structure = StatementStructure.form_structure(statement_title, triggers, cells, titles, decimal, scale_factor,
                                                                      additional_parameters)

    def get_structure(self):
        """Для вызова извне. Считывает структуру таблицы"""
        self._get_structure()
        return self._statement_structure

    def is_template_changed(self) -> bool:
        """ Проверяет, отличается ли в комбо боксе шаблон от вбитого в поля"""
        template = self._statement_structures[self.combo_box.currentText()]
        statement_title, triggers, cells, titles, *__ = StatementStructure.form_output_from_structure(template)

        # curr_statement_title = self.parameter_title.text()
        # curr_triggers = StatementStructure.read_line(self.parameter_trigger.text())
        curr_cells = self.parameter_cells.text()
        curr_titles = self.parameter_column_titles.text()
        # curr_decimal = StatementStructure.read_line(self.parameter_decimal.text())
        # curr_scale_factor = StatementStructure.read_scale(self.scale_factor.text())
        # curr_additional_parameters = self.additional_parameters.text()

        return cells != curr_cells or titles != curr_titles

    @staticmethod
    def form_output_from_structure(structure):
        """
        формирует 5 строк для вывода по ключам из структуры вида:
                structure = {"trigger": ["A"],  #None
                     "columns": {"0": {"title": "Скважина", "cell": "B", "number_of_decimal_places": None, "scale_factor": "*"},
                                 "1": {"title": "Лаб.номер", "cell": "A", "number_of_decimal_places": None, "scale_factor": "*"},
                                 "2": {"title": "Глубина", "cell": "C", "number_of_decimal_places": None, "scale_factor": "*"}}}
        """

        if structure["trigger"] is None:
            structure["trigger"] = [None]

        additional_parameters = structure.get("additional_parameter", "")
        statement_title = structure.get("statement_title", '')

        triggers = ', '.join(str(structure["trigger"][j]) for j in range(len(structure["trigger"])))

        titles = '; '.join(str(structure["columns"][str(j)]["title"]) for j in range(len(structure["columns"])))
        cells = ', '.join(str(structure["columns"][str(j)]["cell"]) for j in range(len(structure["columns"])))
        try:
            decimal = ', '.join(
                str(structure["columns"][str(j)]["number_of_decimal_places"]) for j in range(len(structure["columns"])))
        except:
            decimal = "None"
        try:
            scale_factor = ', '.join(
                str(structure["columns"][str(j)]["scale_factor"]) for j in range(len(structure["columns"])))
        except:
            scale_factor = "*"

        # Пользователю не нужно видеть None
        triggers = triggers.replace(', None', "").replace('None', "")
        cells = cells.replace(', None', "").replace('None', "")
        titles = titles.replace('; None', "").replace('None', "")
        decimal = decimal.replace(', None', "").replace('None', "")


        # scale_factor = scale_factor.replace(', *', '').replace('*', '')

        return statement_title, triggers, cells, titles, decimal, scale_factor, additional_parameters

    @staticmethod
    def read_scale(line):
        line = StatementStructure.read_line(line)
        for i in range(len(line)):
            try:
                float(line[i])
            except:
                if line[i] != '*':
                    line[i] = '*'
        return line

    @staticmethod
    def read_titles(line):

        if line is None:
            s = [None]  # иначе вылетают forы
        else:
            s = [i.strip(" ") for i in line.split(";")]

        return s

    @staticmethod
    def read_ad_params(line):

        if line is None:
            s = [None]  # иначе вылетают forы
        else:
            s = [i.strip(" ") for i in line.split(";")]

        return s

    @staticmethod
    def read_line(line):

        if line is None:
            s = [None]  # иначе вылетают forы
        else:
            s = line.upper().replace(' ', "").split(",")

        return s

    @staticmethod
    def check_lines_len(line1, line2):
        """
        Сравнивает две строки, дополняет меньшую строку до больше через None
        """
        while len(line1) > len(line2):
            line2.append(None)
        while len(line2) < len(line2):
            line2 = line2[:-1]
        return line1, line2

    @staticmethod
    def check_scale_factor_len(line1, line2):
        """
        проверяет длину line2, если она меньше, то в нее дописываются "*"
        """
        while len(line1) > len(line2):
            line2.append("*")
        while len(line1) < len(line2):
            line2 = line2[:-1]
        return line2

    @staticmethod
    def form_structure(statement_title, trigger, cell, title, number_of_decimal_places, scale_factor, additional_parameters):
        """
        Формирует структуру следующего вида
        structure = {"statement_title": "statement_title",
                     "trigger": ["A"],  #None
                     "columns": {"0": {"title": "Скважина", "cell": "B", "number_of_decimal_places": None, "scale_factor": "*"},
                                 "1": {"title": "Лаб.номер", "cell": "A", "number_of_decimal_places": None, "scale_factor": "*"},
                                 "2": {"title": "Глубина", "cell": "C", "number_of_decimal_places": None, "scale_factor": "*"}},
                     "additional_parameter": [additional_parameters]}
        """

        if number_of_decimal_places[0] == "":
            number_of_decimal_places = [None]
        if scale_factor[0] == "":
            scale_factor = ["*"]
        if trigger[0] == "":
            trigger = [None]

        cell, title = StatementStructure.check_lines_len(cell, title)
        cell, number_of_decimal_places = StatementStructure.check_lines_len(cell, number_of_decimal_places)
        scale_factor = StatementStructure.check_scale_factor_len(cell, scale_factor)

        structure = {"statement_title": statement_title,
                     "trigger": trigger,
                     "columns": {str(i): {"title": title[i], "cell": cell[i],
                                          "number_of_decimal_places": number_of_decimal_places[i],
                                          "scale_factor": scale_factor[i]} for i in range(len(cell))},
                     "additional_parameter": additional_parameters}
        return structure

if __name__ == "__main__":
    app = QApplication(sys.argv)

    headlines = ["Лаб. ном.", "Модуль деформации E, кПа", "Сцепление с, МПа",
                 "Угол внутреннего трения, град",
                 "Обжимающее давление 𝜎3", "K0", "Косательное напряжение τ, кПа",
                 "Число циклов N, ед.", "Бальность, балл", "Магнитуда", "Понижающий коэф. rd"]

    fill_keys = ["lab_number", "E", "c", "fi", "sigma3", "K0", "t", "N", "I", "magnituda", "rd"]

    data_test_parameters = {"equipment": ["Выберите прибор", "Прибор: Вилли", "Прибор: Геотек"],
                            "test_type": ["Режим испытания", "Сейсморазжижение", "Штормовое разжижение"],
                            "k0_condition": ["Тип определения K0",
                                             "K0: По ГОСТ-65353", "K0: K0nc из ведомости",
                                             "K0: K0 из ведомости", "K0: Формула Джекки",
                                             "K0: K0 = 1"]
                            }

    Dialog = StatementGenerator(None)
    Dialog.show()
    app.setStyle('Fusion')


    sys.exit(app.exec_())


