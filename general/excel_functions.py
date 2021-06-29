import os
import sys

from general.general_functions import *
from openpyxl import load_workbook
import pyexcel as p
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

from cyclic_loading.cyclic_stress_ratio_function import define_fail_cycle



#Пересохраняем excel
def resave_xls_to_xlsx(file):
    """Пересохраняет файл excel из формата xls в xlsx
            Вернет имя нового файла
            Если файл уже есть, то вернет его"""

    current_file_name = ""

    if file != "":
        # Проверим наличие документа Exel. Если он в старом формате то пересохраним в новый
        if file[-1] == "x":
            current_file_name = file
        elif file[-1] == "s":
            p3 = file + "x"
            if os.path.exists(p3):
                current_file_name = file
                pass
            else:
                p.save_book_as(file_name=file,
                               dest_file_name=p3)  # проверяем есть ли xlsx. Если нет то создаем копию файла в этом формате
            file = p3
            current_file_name = file

        else:
            pass
        current_file_name = file

    return current_file_name



# Считываем с excel
def float_from_excel(s):  # Проверяет строку из Exel и делает ее str. Если она пустая, то возвращает -
    if str(s) == "None":
        return '-'
    else:
        try:
            return float(s)
        except ValueError:
            try:
                return str(s)
            except ValueError:
                return '-'



# Проверки на заполненость
def cell_fullness_test(wb, cell):
    """Проверка заполнения заданной ячейки"""
    cell_data = str(wb["Лист1"][cell].value)
    if cell_data == "None" or cell_data == "-":
        return False
    else:
        return True

def column_fullness_test(wb, columns=[], initial_columns=[]):
    """Проверка заполнения заданной ячейки
    входные параметры:
        columns - список столбцов, которые проверяются на заполненность
        initial_columns - список столбцов, которые будут указателями, какие строки для columns проверять"""

    def test_one_columns(wb, column, initial_columns):
        """Проверка заполнения одного столбца, с учетом initial"""
        all_is_okey = True
        initial_columns.append("A")
        for i in range(7, len(wb['Лист1']['A'])):
            if "None" not in [str(wb["Лист1"][x + str(i)].value) for x in initial_columns]:
                if str(wb["Лист1"][column + str(i)].value) == "None":
                    all_is_okey = False
        return all_is_okey

    for column in columns:
        all_is_okey = test_one_columns(wb, column, initial_columns)
        if all_is_okey == False:
            return False

    return True



# Соответствие ячеек типу испытания
def cfe_test_type_columns(test_type):
    """Функция возвращает столбцы для считывания fi c E по типу испытания"""
    if test_type == "Трёхосное сжатие (E)":
        return ["BI", "BJ", "BK"]

    elif test_type == "Трёхосное сжатие (F, C)":
        return ["BF", "BG", "BH"]

    elif test_type == "Трёхосное сжатие (F, C, E)":
        return ["BC", "BD", "BE"]

    elif test_type == "Трёхосное сжатие с разгрузкой":
        return ["BL", "BM", "BN"]

    elif test_type == "Сейсморазжижение" or test_type == "Штормовое разжижение":
        return ["BZ", "BY", "CA"]

def k0_test_type_column(test_type):
    """Функция возвращает столбцы для считывания K0 по типу испытания"""
    if test_type == "K0: K0nc из ведомости":
        return ["GZ"]
    elif test_type == "K0: K0 из ведомости":
        return ["GY"]
    else: return ["A"]



# Чтение данных
def generator_of_cell_with_lab_number(wb):
    """Функция генерирует последовательность строк с заполненными данными по лабномеру"""
    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None":
            yield i

def currect_lab_number(wb, i):
    """Функция проверяет, заполнен ли новый лабораторный номер в IG"""
    new_lab_number = str(wb["Лист1"]['IG' + str(i)].value)
    if new_lab_number != "None" and new_lab_number != "":
        return new_lab_number
    else:
        return str(wb["Лист1"]['A' + str(i)].value)

def define_c_fi_E(wb, test_mode, i):
    """Отпределяем c fi E"""
    return [float_from_excel(wb["Лист1"][m + str(i)].value) for m in cfe_test_type_columns(test_mode)]

def define_qf(sigma_3, c, fi):
    """Функция определяет qf через обжимающее давление и c fi"""
    fi = fi * np.pi / 180
    return round((2 * (c * 1000 + (np.tan(fi)) * sigma_3)) / (np.cos(fi) - np.tan(fi) + np.sin(fi) * np.tan(fi)), 1)

def define_sigma_3(K0, z):
    """Функция определяет обжимающее давление"""
    return round(K0 * (2 * 9.81 * z), 1)

def define_K0(wb, K0_mode, i, Il, fi):
    """Функция определения K0"""
    if K0_mode == "K0: По ГОСТ-65353":
        if Il == "-":
            return 0.5
        elif Il < 0:
            return 0.6
        elif 0 <= Il < 0.25:
            return 0.7
        elif 0.25 <= Il < 0.5:
            return 0.8
        else:
            return 1
    elif K0_mode == "K0: K0nc из ведомости":
        return round(float_from_excel(wb["Лист1"]['GZ' + str(i)].value), 2)
    elif K0_mode == "K0: K0 из ведомости":
        return round(float_from_excel(wb["Лист1"]['GY' + str(i)].value), 2)
    elif K0_mode == "K0: Формула Джекки":
        return round((1 - np.sin(np.pi * fi / 180)), 2)
    elif K0_mode == "K0: K0 = 1":
        return 1





# Физические свойства
def read_phiz(wb, identification_column=None, identification_color=None):
    """Чтение физических свойств из ведомости
        Передается документ excel, сейсмо или шторм, откуда K0"""
    data = {}

    for i in generator_of_cell_with_lab_number(wb):
        key = currect_lab_number(wb, i)

        try:
            data[key] = read_phiz_line(wb, i, identification_column, identification_color)

        except ValueError:
            pass
    return data

def read_phiz_line(wb, i, identification_column=None, identification_color=None):
    if identification_color and identification_column:
        cell = wb["Лист1"][identification_column + str(i)]
        if cell.font.color != None and type(cell.font.color.rgb) == str:
            if cell.font.color.rgb == "00" + identification_color:
                flag = True
            else:
                flag = False
        else:
            flag = False
    elif identification_column and not identification_color:
        cell = wb["Лист1"][identification_column + str(i)].value
        if cell:
            flag = True
        else:
            flag = False
    else:
        flag = False

    data = {"borehole" : str(wb["Лист1"]['B' + str(i)].value),
                         "depth": float_from_excel(wb["Лист1"]['C' + str(i)].value),
                         "name": float_from_excel(wb["Лист1"]['D' + str(i)].value),
                         "ige": float_from_excel(wb["Лист1"]['ES' + str(i)].value),
                         "rs" : float_from_excel(wb["Лист1"]['P' + str(i)].value),
                         "r": float_from_excel(wb["Лист1"]['Q' + str(i)].value),
                         "rd": float_from_excel(wb["Лист1"]['R' + str(i)].value),
                         "n": float_from_excel(wb["Лист1"]['S' + str(i)].value),
                         "e": float_from_excel(wb["Лист1"]['T' + str(i)].value),
                         "W": float_from_excel(wb["Лист1"]['U' + str(i)].value),
                         "Sr": float_from_excel(wb["Лист1"]['V' + str(i)].value),
                         "Wl": float_from_excel(wb["Лист1"]['W' + str(i)].value),
                         "Wp": float_from_excel(wb["Лист1"]['X' + str(i)].value),
                         "Ip": float_from_excel(wb["Лист1"]['Y' + str(i)].value),
                         "Il": float_from_excel(wb["Лист1"]['Z' + str(i)].value),
                         "Ir": float_from_excel(wb["Лист1"]['AE' + str(i)].value),
                         "str_index": float_from_excel(wb["Лист1"]['AH' + str(i)].value),
                         "gw_depth": float_from_excel(wb["Лист1"]['AJ' + str(i)].value),
                         "build_press": float_from_excel(wb["Лист1"]['AK' + str(i)].value),
                         "pit_depth": float_from_excel(wb["Лист1"]['AL' + str(i)].value),
                         "10": float_from_excel(wb["Лист1"]['E' + str(i)].value),
                         "5": float_from_excel(wb["Лист1"]['F' + str(i)].value),
                         "2": float_from_excel(wb["Лист1"]['G' + str(i)].value),
                         "1": float_from_excel(wb["Лист1"]['H' + str(i)].value),
                         "05": float_from_excel(wb["Лист1"]['I' + str(i)].value),
                         "025": float_from_excel(wb["Лист1"]['J' + str(i)].value),
                         "01": float_from_excel(wb["Лист1"]['K' + str(i)].value),
                         "005": float_from_excel(wb["Лист1"]['L' + str(i)].value),
                         "001": float_from_excel(wb["Лист1"]['M' + str(i)].value),
                         "0002": float_from_excel(wb["Лист1"]['N' + str(i)].value),
                         "0000": float_from_excel(wb["Лист1"]['O' + str(i)].value),
                         "Nop": i,
                         "flag": flag}
    return data

def read_gran(wb, i):
    """Функция считывает грансостав"""
    data = {"10": float_from_excel(wb["Лист1"]['E' + str(i)].value),
            "5": float_from_excel(wb["Лист1"]['F' + str(i)].value),
            "2": float_from_excel(wb["Лист1"]['G' + str(i)].value),
            "1": float_from_excel(wb["Лист1"]['H' + str(i)].value),
            "05": float_from_excel(wb["Лист1"]['I' + str(i)].value),
            "025": float_from_excel(wb["Лист1"]['J' + str(i)].value),
            "01": float_from_excel(wb["Лист1"]['K' + str(i)].value),
            "005": float_from_excel(wb["Лист1"]['L' + str(i)].value),
            "001": float_from_excel(wb["Лист1"]['M' + str(i)].value),
            "0002": float_from_excel(wb["Лист1"]['N' + str(i)].value),
            "0000": float_from_excel(wb["Лист1"]['O' + str(i)].value)}
    return data

# Свойства для трехосника
def read_mech(wb, K0_mode, test_mode = "Трёхосное сжатие (F, C, E)"):
    """Чтение динамических свойств из ведомости"""
    data = {}

    for i in generator_of_cell_with_lab_number(wb):
        key = currect_lab_number(wb, i)

        try:
            c, fi , E = define_c_fi_E(wb, test_mode, i)
            E *= 1000
            if fi != "-" and c != "-" and E != "-":

                # Расчет напряжений
                data_physical = read_phiz_line(wb, i)

                K0 = define_K0(wb, K0_mode, i, data_physical["Il"], fi)

                sigma_3 = define_sigma_3(K0, data_physical["depth"])
                if sigma_3 < 100:
                    sigma_3 =100
                qf = define_qf(sigma_3, c, fi)
                sigma_1 = round(qf + sigma_3, 1)

                poissson = define_poissons_ratio(float_from_excel(wb["Лист1"]['EP' + str(i)].value),
                                               data_physical["Ip"], data_physical["Il"], data_physical["Ir"],
                                               data_physical["10"], data_physical["5"], data_physical["2"])

                Cv = round(float_from_excel(wb["Лист1"]['CC' + str(i)].value), 3)

                Ca = round(float_from_excel(wb["Лист1"]['CF' + str(i)].value), 5)

                build_press = float_from_excel(wb["Лист1"]['AK' + str(i)].value)
                pit_depth = float_from_excel(wb["Лист1"]['AL' + str(i)].value)


                if test_mode == "Трёхосное сжатие с разгрузкой":
                    Eur = round(dependence_E0_Il(data_physical["Il"])*E)
                else:
                    Eur = "-"

                OCR = float_from_excel(wb["Лист1"]['GB' + str(i)].value)
                if OCR == "-":
                    OCR = 1

                dilatancy = round((define_dilatancy_from_xc_qres(define_xc_qf_E(qf, E),
                                                          define_k_q(data_physical["Il"], data_physical["e"],
                                                                     sigma_3)) + define_dilatancy(data_physical,
                                                                                                  data_physical["rs"],
                                                                                                  data_physical["e"],
                                                                                                  sigma_1, sigma_3, fi,
                                                                                                  define_OCR_from_xc(
                                                                                                      define_xc_qf_E(qf,
                                                                                                                     E)),
                                                                                                  data_physical["Ip"],
                                                                                                  data_physical["Ir"]))/2, 2)

                m = define_m(data_physical["e"], data_physical["Il"])
                #m = round(np.random.uniform(0.8, 0.95), 2)
                data[key] = {"E": E, "sigma_3": sigma_3, "sigma_1": sigma_1, "c": c, "fi": fi,
                             "qf": qf, "K0": K0, "Cv": Cv, "Ca": Ca, "poisson": poissson,
                             "build_press": build_press, "pit_depth": pit_depth, "Eur": Eur,
                             "dilatancy": dilatancy, "OCR": OCR, "m": m}

        except ValueError:
            pass
    return data


# Динамические свойства
def read_dynemic(wb, test_mode, K0_mode, ro_mode = "Плотность: 2"):
    """Чтение динамических свойств из ведомости
    Передается документ excel, сейсмо или шторм, откуда K0"""
    Data = {}

    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None":
            if str(wb["Лист1"]['IG' + str(i)].value) != "None":
                key = str(wb["Лист1"]['IG' + str(i)].value)
            else:
                key = str(wb["Лист1"]['A' + str(i)].value)
            try:

                fi = float_from_excel(wb["Лист1"]["BZ" + str(i)].value)
                c = float_from_excel(wb["Лист1"]["BY" + str(i)].value)
                E = float_from_excel(wb["Лист1"]["CA" + str(i)].value) * 1000

                if fi == "-" or c == "-" or E == "-":
                    pass
                else:

                    z = float_from_excel(wb["Лист1"]['C' + str(i)].value)

                    if z <= 9.15:
                        rd = str(round((1 - (0.00765 * z)), 3))
                    elif (z > 9.15) and (z < 23):
                        rd = str(round((1.174 - (0.0267 * z)), 3))
                    else:
                        rd = str(round((1.174 - (0.0267 * 23)), 3))

                    Ip = float_from_excel(wb["Лист1"]['Y' + str(i)].value)
                    Il = float_from_excel(wb["Лист1"]['Z' + str(i)].value)
                    e = float_from_excel(wb["Лист1"]['T' + str(i)].value)

                    if Ip == "-":
                        Ip = 0


                    if K0_mode == "K0: По ГОСТ-65353":
                        if Il == "-":
                            K0 = 0.5
                        elif Il < 0:
                            K0 = 0.6
                        elif 0 <= Il < 0.25:
                            K0 = 0.7
                        elif 0.25 <= Il < 0.5:
                            K0 = 0.8
                        else:
                            K0 = 1
                    elif K0_mode == "K0: K0nc из ведомости":
                        K0 = round(float_from_excel(wb["Лист1"]['GZ' + str(i)].value), 2)
                    elif K0_mode == "K0: K0 из ведомости":
                        K0 = round(float_from_excel(wb["Лист1"]['GY' + str(i)].value), 2)
                    elif K0_mode == "K0: Формула Джекки":
                        K0 = round((1 - np.sin(np.pi * fi / 180)), 2)
                    elif K0_mode == "K0: K0 = 1":
                        K0 = 1
                    else:
                        K0 = 1



                    if float_from_excel(wb["Лист1"]['AJ' + str(i)].value) != "-":
                        Groundwater = float_from_excel(wb["Лист1"]['AJ' + str(i)].value)
                    else:
                        Groundwater = 0

                    if float_from_excel(wb["Лист1"]['Q' + str(i)].value) == "-":
                        ro = 2  # плотность
                        ro_statment = "-"
                    else:
                        ro = float_from_excel(wb["Лист1"]['Q' + str(i)].value)
                        ro_statment = ro

                    if ro_mode == "Плотность: 2":
                        ro = 2
                    elif ro_mode == "Плотность: Из ведомости":
                        pass


                    if test_mode == "Сейсморазжижение":
                        if z <= Groundwater:
                            sig1 = round(ro * 9.81 * z)
                        elif z > Groundwater:
                            sig1 = round(ro * 9.81 * z - (9.81 * (z - Groundwater)))

                        # tau
                        if float_from_excel(wb["Лист1"]['AP' + str(i)].value) != "-":
                            amax = float_from_excel(wb["Лист1"]['AP' + str(i)].value) * 9.81

                            x1 = np.array([0, 0.1 * 9.81, 0.16 * 9.81, 0.24 * 9.81, 0.33 * 9.81, 0.82 * 9.81])
                            y1 = np.array([0, 6, 7, 8, 9, 10])
                            Balnost = np.round(np.interp(amax, x1, y1), 1)

                        else:
                            if float_from_excel(wb["Лист1"]['AM' + str(i)].value) != "-":
                                Balnost = float_from_excel(wb["Лист1"]['AM' + str(i)].value)
                            else:
                                Balnost = 0

                            y1 = np.array([0, 0.1 * 9.81, 0.16 * 9.81, 0.24 * 9.81, 0.33 * 9.81, 0.82 * 9.81])
                            x1 = np.array([0, 6, 7, 8, 9, 10])
                            Ainter = interp1d(x1, y1, kind='cubic')
                            amax = Ainter(Balnost)


                        tau = round(0.65 * amax * (sig1/9.81) * float(rd))
                        if tau < 1:
                            tau = 1

                        # N
                        M = float_from_excel(wb["Лист1"]['AQ' + str(i)].value)
                        if M == "-":
                            M = 0

                        if 0 < M <= 12:
                            y2 = np.array([0, 3, 5, 10, 15, 26, 90])
                            x2 = np.array([0, 5.25, 6, 6.75, 7.5, 8.5, 12])
                            Ninter = interp1d(x2, y2, kind='cubic')
                            N = int(Ninter(M))
                            if N == 0:
                                N = 1
                        else:
                            N = 1
                        try:
                            MSF = round((10 ** (2.24) / ((float(M)) ** (2.56))), 2)
                            tau *= MSF
                            MSF = str(MSF)
                        except ZeroDivisionError:
                            MSF = "-"

                        if sig1 < 10:
                            sig1 = 10

                        if N < 5:
                            N = 5

                        n_fail, Mcsr = define_fail_cycle(N, sig1, tau, Ip, Il, e)

                        Data[key] = {"E": E, "c": c, "fi": fi,
                                     "name": str(wb["Лист1"]['D' + str(i)].value),
                                     "depth": z, "Ip": Ip, "Il": Il, "K0": K0,
                                     "groundwater": Groundwater, "ro": ro_statment,
                                     "balnost": Balnost, "magnituda": M,
                                     "rd": rd, "N": N, "MSF": MSF, "I": Balnost,
                                     "sigma1": sig1, "t": tau, "CSR": round(sig1/tau, 2),
                                     "sigma3": round(sig1*K0),
                                     "frequency": 0.5,
                                     "ige": float_from_excel(
                                         wb["Лист1"]['ES' + str(i)].value),
                                     "n_fail": n_fail, "Mcsr": Mcsr,
                                     "Nop": i}



                    elif test_mode == "Штормовое разжижение":

                        rw = float_from_excel(wb["Лист1"]['HU' + str(i)].value)
                        Hw = float_from_excel(wb["Лист1"]['HS' + str(i)].value)

                        tau = round((0.5 * Hw * rw)/2)

                        sig1 = round((ro - (rw/10)) * 9.81 * z)
                        N = int(float_from_excel(wb["Лист1"]['HR' + str(i)].value))


                        if sig1 < 10:
                            sig1 = 10

                        n_fail, Mcsr = define_fail_cycle(N, sig1, tau, Ip, Il, e)


                        Data[key] = {"E": E, "c": c, "fi": fi,
                                     "name": str(wb["Лист1"]['D' + str(i)].value),
                                     "depth": z, "Ip": Ip, "Il": Il, "K0": K0,
                                     "groundwater": Groundwater, "ro": ro_statment,
                                     "balnost": "-", "magnituda": "-",
                                     "rd": "-", "MSF": "-", "I": "-",
                                     "rw" : rw, "Hw" : Hw, "CSR": round(sig1/tau, 2),
                                     "sigma1": sig1, "t": tau,
                                     "sigma3": round(sig1 * K0),
                                     "n_fail": n_fail, "Mcsr": Mcsr,
                                     "N": N, "frequency": float_from_excel(wb["Лист1"]['HT' + str(i)].value),
                                     "ige": float_from_excel(
                                         wb["Лист1"]['ES' + str(i)].value),
                                     "Nop": i}

            except ValueError:
                pass

    return Data

def read_dynemic_rc(wb, K0_mode, pref_mode):
    """Чтение динамических свойств из ведомости
    Передается документ excel"""
    Data = {}

    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None":

            fi = float_from_excel(wb["Лист1"]["BD" + str(i)].value)
            c = float_from_excel(wb["Лист1"]["BC" + str(i)].value)
            E = float_from_excel(wb["Лист1"]["BE" + str(i)].value)

            if fi == "-":
                pass
            else:

                Ip = float_from_excel(wb["Лист1"]['Y' + str(i)].value)
                Il = float_from_excel(wb["Лист1"]['Z' + str(i)].value)

                if K0_mode == "K0: По ГОСТ-65353":
                    if Ip == "-":
                        K0 = 0.5
                    elif Ip < 1:
                        K0 = 0.5
                    elif Ip > 1 and Il < 0.25:
                        K0 = 0.6
                    elif Ip > 1 and 0.25 <= Il <= 0.75:
                        K0 = 0.7
                    elif Ip > 1 and Il > 0.75:
                        K0 = 0.8
                    else:
                        K0 = 0.5
                elif K0_mode == "K0: K0nc из ведомости":
                    K0 = float_from_excel(wb["Лист1"]['GZ' + str(i)].value)
                elif K0_mode == "K0: K0 из ведомости":
                    K0 = float_from_excel(wb["Лист1"]['GY' + str(i)].value)
                elif K0_mode == "K0: Формула Джекки":
                    K0 = round((1 - np.sin(np.pi * fi / 180)), 2)
                elif K0_mode == "K0: K0 = 1":
                    K0 = 1
                else:
                    K0 = 1

                if pref_mode == "Pref: Pref из столбца FV":
                    pref = round(float_from_excel(wb["Лист1"]['FV' + str(i)].value), 2)
                elif pref_mode == "Pref: Через бытовое давление":
                    pref = round(2 * K0 * float_from_excel(wb["Лист1"]['C' + str(i)].value) * 10/1000, 2)

                Data[str(wb["Лист1"]['A' + str(i)].value)] = {"E": E,
                                                                "c": c,
                                                                "fi": fi,
                                                                "K0": K0,
                                                                "Pref" : pref,
                                                                "Nop": i}

    return Data

# Заказчик
def read_customer(wb):
    """Чтение данных заказчика, даты
        Передается документ excel, возвращает маркер False и данные, либо маркер True и имя ошибки"""

    data = {"customer" : str(wb["Лист1"]["A1"].value),
            "object_name" : str(wb["Лист1"]["A2"].value),
            "data" : str(wb["Лист1"]["Q1"].value),
            "accreditation" : str(wb["Лист1"]["I2"].value),
            "object_number" : str(wb["Лист1"]["AI1"].value)}

    for i in data:

        if data[i] == "None":
            return True, i

    data["data"] = data["data"][:-8]
    return False, data

# Тесты еа заполненность
def ugv_read_test(wb):
    """Проверка заполнения УГВ
    False - все норм"""
    vod = False

    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None" and str(
                wb["Лист1"]["BZ" + str(i)].value) != "None" and str(
            wb["Лист1"]['AJ' + str(i)].value) == "None":
            vod = True

    return vod

def balnost_magnituda_read_test(wb):
    """Проверка заполнения бальности и магнитуды
    False - все норм"""
    bm = False

    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None" and str(wb["Лист1"]["BZ" + str(i)].value) != "None":
            if (str(wb["Лист1"]['AQ' + str(i)].value) == "None"):
                bm = True
                break
            elif (str(wb["Лист1"]['AM' + str(i)].value) == "None" and str(wb["Лист1"]['AP' + str(i)].value) == "None"):
                bm = True
                break

    return bm

def r_water_read_test(wb):
    """Проверка заполнения плотности морской воды
    False - все норм"""
    vod = False

    if str(wb["Лист1"]['HU7'].value) != "None":
        vod = True

    return vod

def pref_read_test(wb):
    """Проверка заполнения pref
    False - все норм"""
    pref = False

    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None" and str(
            wb["Лист1"]['FV' + str(i)].value) == "None":
            pref = True

    return pref

def storm_read_test(wb):
    """Проверка заполнения УГВ
    False - все норм"""
    data = False

    for i in range(7, len(wb['Лист1']['A'])):
        if str(wb["Лист1"]['A' + str(i)].value) != "None" \
                and str(wb["Лист1"]["BZ" + str(i)].value) != "None" \
                and (str(wb["Лист1"]['HR' + str(i)].value) == "None"\
                or str(wb["Лист1"]['HS' + str(i)].value) == "None"\
                or str(wb["Лист1"]['HT' + str(i)].value) == "None"\
                or str(wb["Лист1"]['HU' + str(i)].value) == "None"):
            data = True
    return data

def K0_read_test(wb, mode):
    """ЧПроверка заполнености K0 в ведомости
            Falsr - все норм"""

    k0 = False
    for i in range(7, len(wb['Лист1']['A'])):

        if mode == "K0: K0 из ведомости":
            if str(wb["Лист1"]['A' + str(i)].value) != "None" and str(
                    wb["Лист1"]["BZ" + str(i)].value) != "None" and str(
                wb["Лист1"]['GY' + str(i)].value) == "None":
                k0 = True
        elif mode == "K0: K0nc из ведомости":
            if str(wb["Лист1"]['A' + str(i)].value) != "None" and str(
                    wb["Лист1"]["BZ" + str(i)].value) != "None" and str(
                wb["Лист1"]['GZ' + str(i)].value) == "None":
                k0 = True

    return k0




# пишем в excel
def write_cyclic_result_to_excel(path, Lab, Eps, PPR):
    """Запись в файл excel"""
    wb = load_workbook(path)

    iLab_new = 0
    iLab = 0

    for i in range(7, len(wb['Лист1']['A'])):

        if str(wb["Лист1"]['IG' + str(i)].value) == Lab:
            iLab_new = i
        elif str(wb["Лист1"]['A' + str(i)].value) == Lab:
            iLab = i

    if iLab_new != 0:
        wb["Лист1"]['HX' + str(iLab_new)] = PPR
        wb["Лист1"]['HW' + str(iLab_new)] = Eps
    elif iLab != 0:
        wb["Лист1"]['HX' + str(iLab)] = PPR
        wb["Лист1"]['HW' + str(iLab)] = Eps

    wb.save(path)

def set_cell_data(path, cell, value, sheet="Лист1", color=None)->None:
    """Запись в файл excel
    :argument
        lab_number->str: Лабораторный номер
        column->str: Ячейка для записи
        value: Записываемое значение
        sheet->str: Лист для записи
        color->str: цвет шрифта записи
    :return
        None"""

    wb = load_workbook(path)
    wb[sheet][cell] = value
    if color:
        cell = wb[sheet][cell]
        cell.font = Font(color=color)
    wb.save(path)

def write_to_excel(path, Lab, params):
    """Запись в файл excel"""
    wb = load_workbook(path)

    wb["Лист1"]["HY5"] = "Сигма1, кПа"
    wb["Лист1"]["HZ5"] = "Сигма3, кПа"
    wb["Лист1"]["IA5"] = "Тау, кПа"
    wb["Лист1"]["IB5"] = "K0"
    wb["Лист1"]["IC5"] = "Частота, Гц"
    wb["Лист1"]["ID5"] = "цикл разрушения"

    iLab_new = 0
    iLab = 0

    for i in range(7, len(wb['Лист1']['A'])):

        if str(wb["Лист1"]['IG' + str(i)].value) == Lab:
            iLab_new = i
        elif str(wb["Лист1"]['A' + str(i)].value) == Lab:
            iLab = i

    if iLab_new != 0:
        cells = ["HW", "HX", "HY", "HZ", "IA", "IB", "IC", "ID", "IE", "IF"]
        for param, cell in zip(params, cells):
            wb["Лист1"][cell + str(iLab_new)] = param

    elif iLab != 0:
        cells = ["HW", "HX", "HY", "HZ", "IA", "IB", "IC", "ID", "IE", "IF"]
        for param, cell in zip(params, cells):
            wb["Лист1"][cell + str(iLab)] = param

    wb.save(path)

def write_rc_result_to_excel(path, Lab, G0, gam):
    """Запись в файл excel"""
    wb = load_workbook(path)


    iLab_new = 0
    iLab = 0

    for i in range(7, len(wb['Лист1']['A'])):

        if str(wb["Лист1"]['IG' + str(i)].value) == Lab:
            iLab_new = i
        elif str(wb["Лист1"]['A' + str(i)].value) == Lab:
            iLab = i

    if iLab_new != 0:
        wb["Лист1"]['HL' + str(iLab_new)] = G0
        wb["Лист1"]['HK' + str(iLab_new)] = gam
    elif iLab != 0:
        wb["Лист1"]['HL' + str(iLab)] = G0
        wb["Лист1"]['HK' + str(iLab)] = gam

    wb.save(path)



def get_column_letters(last_letter='IV'):
    """
    Функция формирует список наименований колонок из exel (ключей)
    :param last_letter: str, название колонки по которую (включительно) формировать словарь
    :return: list, список наименований колонок (ключи)
    """
    import_columns = []
    try:
        last_letter_index = column_index_from_string(last_letter)
    except:
        last_letter_index = column_index_from_string('IV')
    for i in range(1, last_letter_index + 1):
        import_columns.append(get_column_letter(i))
    return import_columns[:last_letter_index + 1]  # обрезание идет До индекса, так что включаем еще значение


def table_data(table, structure):
    """Функция возвращает матрицу для построения таблицы. Первая втрока - имя, остальные - столбцы значений
    Входные параметры: table - матрица, считанная с excel,
                       structure - словарь, описывающий структуру таблицы
                       structure = {"trigger": ['BI', 'AS'],
                       "columns": {"0": {"title": "Скважина", "cell": "B"},
                                  "1": {"title": "Лаб.номер", "cell": "A"},
                                  "2": {"title": "Глубина", "cell": "C"}}}"""

    data = [[]]

    titles = [structure["columns"][str(i)]["title"] for i in range(len(structure["columns"]))]

    scale = [structure["columns"][str(i)]["scale_factor"] for i in range(len(structure["columns"]))]

    parameter_decimal = [structure["columns"][str(i)]["number_of_decimal_places"] for i in
                         range(len(structure["columns"]))]

    # for i in range(len(structure["columns"])): # идем по строкам columns
    #     data[0].append(structure["columns"][str(i)]["title"])  # в список списков в первый список записываем все title из columns str(i)-дает обращение к нужному ключу по порядку

    if (structure["trigger"] is None) or (structure["trigger"] == []):
        structure["trigger"] = [None]
    while len(structure["trigger"]) > 1 and structure["trigger"].count(None) > 0:
        structure["trigger"].remove(
            None)  # удаляем None так, чтобы остался массив из одного None на случай массива [None, A]

    if structure["trigger"].count(None) == 0:
        k = 0
        for i in range(len(table[structure["trigger"][0]])):  # Идем по столбцу тригера
            flag = 1
            for tr in range(len(structure["trigger"])):  # внутренний цикл для того чтобы идти по всем тригерам
                if table[structure["trigger"][tr]][
                    i] == 'None':  # если в ячейке из столбца тригера пусто меняем флаг на 0
                    flag = 0
            if flag:  # для непустых ячеек тригера дбавляем в список списков пустые списки
                for j in range(len(structure["columns"])):  # по длине массива с названими
                    data[k].append(table[structure["columns"][str(j)]["cell"]][
                                       i])  # записываем в каждый массив значения из каждой строки в нужных столбцах
                data.append([])
                k += 1  # считаем количество строк которые были записаны
        data.pop(-1)
    else:  # если тригера нет
        for i in range(len(table[structure["columns"]["0"]["cell"]])):  # идем по столбцу первому из columns
            for j in range(len(structure["columns"])):
                data[i].append(table[structure["columns"][str(j)]["cell"]][
                                   i])  # записываем в каждый массив значения из каждой строки в нужных столбцах
            data.append([])  # добавляем нужное число пустых массивов
        data.pop(-1)
    # print('до округ',data)
    # Форматирование данных под число знаком после запятой
    data = number_of_decimal_places(data, parameter_decimal)

    # Перевод scale к виду scales = [3 * cm, 5 * cm, 6 * cm]
    for i in range(len(scale)):
        try:
            scale[i] = 28.346456692913385*float(scale[i]) # значение cm в from reportlab.lib.units import inch, cm, mm
        except ValueError:
            pass
    s = [i.strip(" ") for i in structure["additional_parameter"].split(";")]
    for i in range(len(s)):
        data.append([s[i]])
    #data.append([structure["additional_parameter"]])

    # print('после округ',data)
    return titles, data, scale


def form_xlsx_dictionary(wb, last_key):
    """
    Функция считывает всю ведомость и записывает значения в словарь, где
    ключи - обозначения колонок в exel таблице ('A', 'B', ...)
    значения - колонок из ведомости в виде массивов
    :param wb: workbook, результат импорта ведомости
    :param last_key: str, название колонки (ключа) по которую (включительно) формировать словарь
    :return: dict, словарь с ключами по наименованиям колонок и соответствующими массивами колонок - numpy.ndarray
    """

    # наименования колонок большими буквами
    last_key = last_key.upper()

    # формируем список ключей
    import_columns = get_column_letters(last_key)

    # объявляем словарь
    xlsx_dictionary = {str(import_columns[0]): []}
    # вносим в него ключи
    for col in import_columns:
        xlsx_dictionary[str(col)] = []

    for key in xlsx_dictionary:
        for i in generator_of_cell_with_lab_number(wb):
            # выполняем проверку, что ячейка не пустая
            if str(wb["Лист1"][str(key) + str(i)].value) != "None":
                xlsx_dictionary[str(key)].append(wb["Лист1"][str(key) + str(i)].value)
            else:
                xlsx_dictionary[str(key)].append("None")
    # переводим значения в массивы numpy.ndarray
    for key in xlsx_dictionary:
        xlsx_dictionary[str(key)] = np.array(xlsx_dictionary[str(key)])

    return xlsx_dictionary




def number_of_decimal_places(matrix, parameter_decimal):
    for j, count in enumerate(parameter_decimal):
        if str(count) != 'None' and str(count) != '*' and str(count) != '':
            for i in range(len(matrix)):
                try:
                    matrix[i][j] = number_format(float_float(matrix[i][j]), characters_number=int(count), split='.')

                except:
                    pass

    # !!!
    for i in range(len(matrix)):
        for j in range(len(matrix[0])):
            try:
                matrix[i][j] = str(matrix[i][j])
            except:
                pass

    return matrix

def float_float(a):
    try:
        a = float(a)
        return a
    except ValueError:
        try:
            a = float(a.replace(",", "."))
            return a
        except ValueError:
            return a


if __name__ == "__main__":
    amax = 0.5
    x1 = np.array([0, 0.1 * 9.81, 0.16 * 9.81, 0.24 * 9.81, 0.33 * 9.81, 0.82 * 9.81])
    y1 = np.array([0, 6, 7, 8, 9, 10])
    Ainter = interp1d(x1, y1, kind='linear')
    Balnost = Ainter(0.5 * 9.81)
    Balnost = np.round(Balnost, 3)


    print(np.round(np.interp(amax* 9.81, x1, y1), 1))

    import matplotlib.pyplot as plt
    plt.plot(x1, y1)
    xxx = np.linspace(0, 0.82 * 9.81, 100)
    plt.plot(xxx, Ainter(xxx))
    plt.show()
    print(Balnost)