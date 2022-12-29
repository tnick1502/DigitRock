from copy import copy
from typing import List, Union

import openpyxl
from openpyexcel.utils import get_column_letter, coordinate_from_string, column_index_from_string
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment


# file = r"C:\\reports\\test_join.xlsx"
# img = r"C:\\reports\\logo.png"
#
# wb = openpyxl.load_workbook(file)
# ws = wb.worksheets[0]
# # img = openpyxl.drawing.image.Image(img)
# # img.anchor = 'A1'
# # ws.add_image(img)
# wb.save('C:\\reports\\out.xlsx')


class ReportXlsxSaver:
    """
        Класс для сохранения ведомостьей в xls формате.
        Поддерживает два режима и два типа обработки данных.

        Режимы:
            1. Режим шаблона `__template_mode`. Выставляется в true, если в инит передан
            пользовательский шаблон в виде `template_filename`
            В этом режиме НЕ осуществляется печать заголовков столбцов и
            НЕ осуществляется вставка новых строк. Число строк шаблона передается через num_page_rows в ините

            2. Режим стандартного шаблона, выбирается если название шаблона не передавалось.
            Тогда названия столбцов заполняеются по ключам из `tests_data`

        Типы обработки:
            1. _dict_mode - Режим словаря (ТОЛЬКО ДЛЯ ОДНОЙ СТРАНИЦЫ)
                Функция для подготовки данных - form_tests_data_dict_mode
                Фукнция для печати - set_data_dict_mode
            В этом случае данные должны лежать в словаре внутри tests_data по ключам в виде имен столбцов

            2. _list_mode - Построчный режим (ТОЛЬКО ДЛЯ ШАБЛОННОГО РЕЖИМА)
                Функция для подготовки данных - form_tests_data_list_mode
                Фукнция для печати - set_data_row_mode
            В этом случае данные заполняеются построчно, названий столцов нет, они должны быть заданы в шаблоне
    """
    def __init__(self, template_filename: str = None, num_page_rows: int = 55):
        self.__TEMPLATE_PATH: str = r"general/xls_statment_base_template.xlsx"
        '''относительный путь до шаблона'''

        self.__template_mode: bool = False
        '''режим, при котором выбран конкретный шаблон,
         в этом режиме пропускаются некоторые фукнции,
          например, вставка строк и заполнение заголовков'''

        self.page_size_rows: int = 55
        '''число строчек в шаблоне'''

        self.page_slant: int = 0
        '''текущий сдвиг по листам, увеличивается на `page_size_rows` для каждого листа'''

        if template_filename:
            self.__template_filename = template_filename
            '''имя файла шаблона'''

            self.__TEMPLATE_PATH = f'{self.__TEMPLATE_PATH.split("/")[0]}/{template_filename}'
            self.page_size_rows = num_page_rows
            self.__template_mode = True

        self.__book: 'Workbook' = None
        self.__sheet: 'Worksheet' = None
        self.__tests_data: 'dict' = {'labNum': {'skv': None, 'depth': None, 'type': None, 'params': {}}}
        self.__params_positions: 'dict' = {}
        self._base_height = 15

        self.__book = openpyxl.load_workbook(self.__TEMPLATE_PATH)
        self.__sheet = self.__book.worksheets[0]

    def set_data_dict_mode(self, customer: 'str', obj_name: 'str', test_title: 'str', date: 'str', tests_data: 'dict',
                           additional_data: 'List' = None, accreditation: 'str' = None, doc_num: 'str' = None):
        """

        """
        self.__tests_data = copy(tests_data)

        if len(self.__tests_data.keys()) < 1:
            return

        # данные по объекту
        self.__sheet[self.pos_of('date')].value = date
        self.__sheet[self.pos_of('customer')].value = customer
        self.__sheet[self.pos_of('obj_name')].value = obj_name
        self.__sheet[self.pos_of('test_title')].value = test_title
        if doc_num:
            self.__sheet[self.pos_of('doc_num')].value = doc_num

        # дополнителньые данные
        if accreditation:
            accreditation_font = copy(self.__sheet[self.pos_of('accreditation')].font)
            accreditation_border = copy(self.__sheet[self.pos_of('accreditation')].border)
            accreditation_fill = copy(self.__sheet[self.pos_of('accreditation')].fill)

            self.__sheet[self.pos_of('accreditation')].value = accreditation
            self.__sheet[self.pos_of('accreditation')].font = accreditation_font
            self.__sheet[self.pos_of('accreditation')].border = accreditation_border
            self.__sheet[self.pos_of('accreditation')].fill = accreditation_fill
            self.__sheet[self.pos_of('accreditation')].alignment = Alignment(wrapText=True,
                                                                             horizontal='center',
                                                                             vertical='center')

        # Заполняем заголовки с параметрами опыта:
        _first_col = self.pos_of('last_title_col') + 1
        _params = [*self.__tests_data.keys()]
        if _params and len(_params) > 0:
            for param in self.__tests_data[[*self.__tests_data.keys()][0]]['params']:
                self.__params_positions[param] = get_column_letter(_first_col)

                if not self.__template_mode:
                    self.set_cell_value(f"{self.__params_positions[param]}{self.pos_of('first_row')}", value=param,
                                        style_key='labNum')
                    lab_num_width = self.__sheet.column_dimensions[self.pos_of('labNum')].width
                    self.__sheet.column_dimensions[self.__params_positions[param]].width = lab_num_width

                _first_col = _first_col + 1

        # Начинаем вставлять строчки с опытами
        _cur_row = self.pos_of('first_row') + 1  # строка с заголовками + 1
        for labnum in self.__tests_data:
            # Вставляем строчку с новым опытом и записываем лобораторный номер
            if not self.__template_mode:
                self.__sheet.insert_rows(idx=_cur_row)
            self.__sheet.row_dimensions[_cur_row].height = self._base_height
            self.set_cell_value(self.pos_of('labNum')+str(_cur_row), value=labnum, style_key='labNum')
            # Далее проходим по всем свойствам пробы кроме дополнительных 'params'
            for prop in self.__tests_data[labnum]:
                if prop != 'params':
                    self.set_cell_value(f"{self.pos_of(prop)}{_cur_row}",
                                        value=self.__tests_data[labnum][prop],
                                        style_key=prop)
                    if prop == 'type':
                        self.__sheet.merge_cells(f"{self.pos_of('type')}{_cur_row}"
                                                 f":{get_column_letter(self.pos_of('last_title_col'))}{_cur_row}")
                    continue

                # Параметры заполняем отдельно для каждого цикла
                for param in self.__tests_data[labnum]['params']:
                    self.set_cell_value(f"{self.__params_positions[param]}{_cur_row}",
                                        value=self.__tests_data[labnum]['params'][param],
                                        style_key='depth')
            # Конец заполнения свойств пробы
            _cur_row = _cur_row + 1

        # После всех проб заполняем дополнительные параметры
        _last_col = get_column_letter(self.pos_of('last_title_col'))
        if self.__params_positions:
            _values = [*self.__params_positions.values()]
            _last_col = max(_values)
        self.__sheet.merge_cells(f"{self.pos_of('labNum')}{_cur_row}:{_last_col}{_cur_row}")
        _cur_row = _cur_row + 1
        if additional_data:
            for data in additional_data:
                if not self.__template_mode:
                    self.__sheet.insert_rows(idx=_cur_row)
                self.__sheet.row_dimensions[_cur_row].height = self._base_height
                self.set_cell_value(f"{self.pos_of('labNum')}{_cur_row}",
                                    value=data,
                                    style_key='labNum')
                self.__sheet.merge_cells(f"{self.pos_of('labNum')}{_cur_row}:{_last_col}{_cur_row}")
                _cur_row = _cur_row + 1
        # Конец заполнения дополнительных параметров

        _b_table_first_row = coordinate_from_string(self.pos_of('date'))[-1] + (_cur_row - self.pos_of('first_row')) - 3
        '''первая строчка нижней таблицы, + 2 дает последнюю строку листа'''

        # Добавляем строки так, чтобы документ был на весю высоту А4, если строк недостаточно
        if not self.__template_mode:
            if (_b_table_first_row + 2) < self.page_size_rows:
                _size_correction = self.page_size_rows - (_b_table_first_row + 2)
                self.__sheet.insert_rows(idx=_cur_row + 1, amount=_size_correction)
                _b_table_first_row += _size_correction

        # Теперь нужно привести в порядок табличку внизу отчета
        if not self.__template_mode:
            self.__sheet.row_dimensions[_b_table_first_row - 6].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row - 5].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row - 4].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row + 1].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row + 2].height = self._base_height

            self.__sheet.merge_cells(f"{get_column_letter(self.pos_of('bottomTableJoinLastCol')+1)}{_b_table_first_row + 1}:"
                                     f"{get_column_letter(self.pos_of('bottomTableJoinLastCol')+1)}{_b_table_first_row + 2}")

            self.__sheet.merge_cells(f"{self.pos_of('bottomTableJoinFirstCol')}{_b_table_first_row}:"
                                     f"{get_column_letter(self.pos_of('bottomTableJoinLastCol'))}{_b_table_first_row + 2}")

            # Вся экселька обернута в границы, поэтому необходимо
            left_border = copy(self.__sheet['S5'].border)
            right_border = copy(self.__sheet['A5'].border)
            for i in range(2, _b_table_first_row):
                self.__sheet[f"S{i}"].border = left_border
                self.__sheet[f"A{i}"].border = right_border

            self.__sheet.print_area = f"A1:{get_column_letter(self.pos_of('bottomTableJoinLastCol') + 1)}" \
                                      f"{_b_table_first_row + 2}"

    def set_data_row_mode(self, customer: 'str', obj_name: 'str', test_title: 'str', date: 'str', tests_data: 'dict',
                          additional_data: 'List' = None, accreditation: 'str' = None, doc_num: 'str' = None):

        _last_page_num = 1
        for page in range(len(tests_data) - 1):
            self.print_page_row_mode(customer, obj_name, test_title, date, tests_data[page],
                                     additional_data=None, page_of_pages=f'{page + 1}/{len(tests_data)}', doc_num=doc_num)
            self.page_slant += self.page_size_rows
            _last_page_num += 1

        try:
            col_count = len(tests_data[0][0])
        except IndexError:
            col_count = None

        self.print_page_row_mode(customer, obj_name, test_title, date, tests_data[-1], additional_data, accreditation,
                                 page_of_pages=f'{_last_page_num}/{len(tests_data)}', doc_num=doc_num,
                                 col_count=col_count)

    def print_page_row_mode(self, customer: 'str', obj_name: 'str', test_title: 'str', date: 'str', tests_data: 'List',
                            additional_data: 'List' = None, accreditation: 'str' = None, page_of_pages: 'str' = '1/1',
                            doc_num: 'str' = None, col_count: 'int' = None):
        """

        """
        self.__tests_data = copy(tests_data)

        # данные по объекту
        self.__sheet[self.pos_of('date')].value = date
        self.__sheet[self.pos_of('customer')].value = customer
        self.__sheet[self.pos_of('obj_name')].value = obj_name
        self.__sheet[self.pos_of('test_title')].value = test_title
        self.__sheet[self.pos_of('page_of_pages')].value = page_of_pages
        if doc_num:
            self.__sheet[self.pos_of('doc_num')].value = doc_num

        # дополнителньые данные
        if accreditation:
            accreditation_font = copy(self.__sheet[self.pos_of('accreditation')].font)
            accreditation_border = copy(self.__sheet[self.pos_of('accreditation')].border)
            accreditation_fill = copy(self.__sheet[self.pos_of('accreditation')].fill)

            self.__sheet[self.pos_of('accreditation')].value = accreditation
            self.__sheet[self.pos_of('accreditation')].font = accreditation_font
            self.__sheet[self.pos_of('accreditation')].border = accreditation_border
            self.__sheet[self.pos_of('accreditation')].fill = accreditation_fill
            self.__sheet[self.pos_of('accreditation')].alignment = Alignment(wrapText=True,
                                                                             horizontal='center',
                                                                             vertical='center')

        # Начинаем вставлять строчки с опытами
        _cur_row = self.pos_of('first_row') + 1  # строка с заголовками + 1
        for row in self.__tests_data:
            # Вставляем строчку с новым опытом и записываем лобораторный номер
            if not self.__template_mode:
                self.__sheet.insert_rows(idx=_cur_row)
            self.__sheet.row_dimensions[_cur_row].height = self._base_height
            self.set_cell_value(self.pos_of('labNum')+str(_cur_row), value=row[0], style_key='labNum')
            self.set_cell_value(self.pos_of('skv')+str(_cur_row), value=row[1], style_key='skv')
            self.set_cell_value(self.pos_of('depth')+str(_cur_row), value=row[2], style_key='depth')
            self.set_cell_value(self.pos_of('type')+str(_cur_row), value=row[3], style_key='type')
            self.__sheet.merge_cells(f"{self.pos_of('type')}{_cur_row}"
                                     f":{get_column_letter(self.pos_of('last_title_col'))}{_cur_row}")
            # Далее проходим по всем свойствам пробы кроме дополнительных 'params'
            for i in range(4, len(row)):
                self.set_cell_value(f"{get_column_letter(self.pos_of('last_title_col') + i - 3)}{_cur_row}",
                                    value=row[i],
                                    style_key='depth')
            # Конец заполнения свойств пробы
            _cur_row = _cur_row + 1

            if _cur_row > self.page_size_rows - 10 + self.page_slant:
                break

        # После всех проб заполняем дополнительные параметры
        if col_count is None:
            if len(self.__tests_data) == 0:
                col_count = self.pos_of('bottomTableJoinLastCol') - self.pos_of('last_title_col') + 4 - 1
            else:
                col_count = len(self.__tests_data[0])
        _last_col = get_column_letter(col_count + self.pos_of('last_title_col') - 4)

        # self.__sheet.merge_cells(f"{self.pos_of('labNum')}{_cur_row}:{_last_col}{_cur_row}")
        # _cur_row = _cur_row + 1

        if additional_data:
            for data in additional_data:
                if not self.__template_mode:
                    self.__sheet.insert_rows(idx=_cur_row)
                self.__sheet.row_dimensions[_cur_row].height = self._base_height
                self.set_cell_value(f"{self.pos_of('labNum')}{_cur_row}",
                                    value=data,
                                    style_key='labNum')
                self.__sheet.merge_cells(f"{self.pos_of('labNum')}{_cur_row}:{_last_col}{_cur_row}")
                _cur_row = _cur_row + 1
        # Конец заполнения дополнительных параметров

        _b_table_first_row = coordinate_from_string(self.pos_of('date'))[-1] + (_cur_row - self.pos_of('first_row')) - 2
        '''первая строчка нижней таблицы, + 2 дает последнюю строку листа'''

        # Добавляем строки так, чтобы документ был на весю высоту А4, если строк недостаточно
        if not self.__template_mode:
            if (_b_table_first_row + 2) < self.page_size_rows:
                _size_correction = self.page_size_rows - (_b_table_first_row + 2)
                self.__sheet.insert_rows(idx=_cur_row + 1, amount=_size_correction)
                _b_table_first_row += _size_correction

        # Теперь нужно привести в порядок табличку внизу отчета
        if not self.__template_mode:
            self.__sheet.row_dimensions[_b_table_first_row - 6].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row - 5].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row - 4].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row + 1].height = self._base_height
            self.__sheet.row_dimensions[_b_table_first_row + 2].height = self._base_height

            self.__sheet.merge_cells(f"{get_column_letter(self.pos_of('bottomTableJoinLastCol')+1)}{_b_table_first_row + 1}:"
                                     f"{get_column_letter(self.pos_of('bottomTableJoinLastCol')+1)}{_b_table_first_row + 2}")

            self.__sheet.merge_cells(f"{self.pos_of('bottomTableJoinFirstCol')}{_b_table_first_row}:"
                                     f"{get_column_letter(self.pos_of('bottomTableJoinLastCol'))}{_b_table_first_row + 2}")

            # Вся экселька обернута в границы, поэтому необходимо
            left_border = copy(self.__sheet['U5'].border)
            right_border = copy(self.__sheet['A5'].border)
            for i in range(2, _b_table_first_row):
                self.__sheet[f"U{i}"].border = left_border
                self.__sheet[f"A{i}"].border = right_border

            self.__sheet.print_area = f"A1:{get_column_letter(self.pos_of('bottomTableJoinLastCol') + 1)}" \
                                      f"{_b_table_first_row + 2}"

    def save(self, path):
        self.__book.save(path)

    def set_cell_value(self, cell, value, style_key):
        style_cell = f"{self.pos_of(style_key)}{self.pos_of('first_row')}"
        self.__sheet[cell].value = value
        self.__sheet[cell].font = copy(self.__sheet[style_cell].font)
        self.__sheet[cell].border = copy(self.__sheet[style_cell].border)
        self.__sheet[cell].fill = copy(self.__sheet[style_cell].fill)
        self.__sheet[cell].alignment = copy(self.__sheet[style_cell].alignment)

        if type(value) in (int, float):
            split = str(value).split('.')
            if len(split) == 1:
                self.__sheet[cell].number_format = '0.00'
                return
            if len(split) > 1:
                strlen = len(str(value).split('.')[-1])
                self.__sheet[cell].number_format = f"0.{'0'*strlen}"

    def pos_of(self, key: str) -> Union[str, int]:
        """
        Функция для опередения прибитых позиций в шаблоне.
        Возвращает или координаты ячеки в str или номер колонки или столбца,
         в зависимости от поданного ключа.
        Позиции в шаблоне выбирает исходя из заданных параметров шаблона:
         `self.__template_mode` и `self.__template_filename`
        Строки корректируются на сдвиг текущего листа `self.page_slant`
        """

        if not self.__template_mode:
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{54 + self.page_slant}', 'doc_num': f'H{54 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 18,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 12,
                           'page_of_pages': f'Q{54 + self.page_slant}'}
            return __positions[key]

        if self.__template_filename == 'xls_statment_VIBRO_template.xlsx':
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{54 + self.page_slant}', 'doc_num': f'H{54 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 12,
                           'page_of_pages': f'Q{54 + self.page_slant}'}
            return __positions[key]

        if self.__template_filename == 'xls_statment_DEMPH_template.xlsx':
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{53 + self.page_slant}', 'doc_num': f'H{53 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 13,
                           'page_of_pages': f'Q{53 + self.page_slant}'}
            return __positions[key]

        if self.__template_filename == 'xls_statment_RELEY_template.xlsx':
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{53 + self.page_slant}', 'doc_num': f'H{53 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 11,
                           'page_of_pages': f'Q{53 + self.page_slant}'}
            return __positions[key]

        if self.__template_filename == 'xls_statment_RESONANT_template.xlsx':
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{53 + self.page_slant}', 'doc_num': f'H{53 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 13,
                           'page_of_pages': f'Q{53 + self.page_slant}'}
            return __positions[key]

        if self.__template_filename == 'xls_statment_SEISMO_template.xlsx':
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{52 + self.page_slant}', 'doc_num': f'H{52 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 11,
                           'page_of_pages': f'R{52 + self.page_slant}'}
            return __positions[key]

        if self.__template_filename == 'xls_statment_STORM_template.xlsx':
            __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                           'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                           'date': f'J{52 + self.page_slant}', 'doc_num': f'H{52 + self.page_slant}',
                           'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                           'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                           'first_row': 13 + self.page_slant, 'last_title_col': 11,
                           'page_of_pages': f'R{52 + self.page_slant}'}
            return __positions[key]

        __positions = {'customer': f'L{2 + self.page_slant}', 'obj_name': f'L{4 + self.page_slant}',
                       'accreditation': f'C{7}', 'test_title': f'B{11 + self.page_slant}',
                       'date': f'J{54 + self.page_slant}', 'doc_num': f'H{54 + self.page_slant}',
                       'bottomTableJoinFirstCol': 'K', 'bottomTableJoinLastCol': 17,
                       'labNum': 'B', 'skv': 'C', 'depth': 'D', 'type': 'E',
                       'first_row': 13 + self.page_slant, 'last_title_col': 12,
                       'page_of_pages': f'Q{54 + self.page_slant}'}
        return __positions[key]

    def form_tests_data_list_mode(self, titles, data):
        """
        titles = ['Лаб. №', 'Скв. №', 'Глубина отбора, м', 'Наименование грунта', '<p>Референтное давление p<sub rise="0.5" size="5">ref</sub>]
        data = [['8-1', '8.0', '28,8', 'Супесь пластичная', '0,350', '100,8', '3,58'], ['']]
        """
        if len(titles) < 4:
            return {}

        _result = [[]]
        _additional = []
        _page = 0

        for ind, row in enumerate(data):
            if ind > self.page_size_rows - 10 - (self.pos_of("first_row") + self.page_slant) + (_page * self.page_size_rows):
                _page += 1
                _result.append([])

            if len(row) == 1 and len(row[0]) > 0:
                _additional.append(row[0])
                continue
            if len(row) < 4:
                continue

            _result[_page].append(row)

        return _result, _additional

    @staticmethod
    def form_tests_data_dict_mode(titles, data):
        """
        titles = ['Лаб. №', 'Скв. №', 'Глубина отбора, м', 'Наименование грунта', '<p>Референтное давление p<sub rise="0.5" size="5">ref</sub>]
        data = [['8-1', '8.0', '28,8', 'Супесь пластичная', '0,350', '100,8', '3,58'], ['']]
        """
        if len(titles) < 4:
            return {}

        _result = {}
        _additional = []

        for ind, row in enumerate(data):
            if len(row) == 1 and len(row[0]) > 0:
                _additional.append(row[0])
                continue
            if len(row) < 4:
                continue

            _result[row[0]] = {'skv': row[1], 'depth': row[2], 'type': row[3]}

            if 'params' not in _result[row[0]]:
                _result[row[0]]['params'] = {}

            for i in range(4, len(row)):
                if not titles[i]:
                    titles[i] = ''
                title = titles[i].replace('<p>', '').replace('</p>', '').replace("<sub>", '')\
                    .replace('</sub>', '').replace('<sub rise="0.5" size="5">', '')\
                    .replace('<sup rise="-2" size="4">', 'E').replace('</sup>', '')\
                    .replace(' *10', '').replace('*10', '')
                title = title.replace('&alpha', 'α').replace('&beta', 'β').replace('&gamma', 'γ')\
                    .replace('&delta', 'δ')\
                    .replace('&epsilon', 'ε').replace('&zeta', 'ζ').replace('&eta', 'η')\
                    .replace('&theta', 'θ').replace('&iota', 'ι').replace('&kappa', 'κ')\
                    .replace('&lambda', 'λ').replace('&mu', 'μ').replace('&nu', 'ν')\
                    .replace('&xi', 'ξ').replace('&omikron', 'ο').replace('&pi', 'π') \
                    .replace('&rho', 'ρ').replace('&sigma', 'σ').replace('&tau', 'τ') \
                    .replace('&upsilon', 'υ').replace('&phi', 'φ').replace('&chi', 'χ') \
                    .replace('&psi', 'ψ').replace('&omega', 'ω')
                _result[row[0]]['params'][title] = row[i]

        return _result, _additional



def test_report():
    file = r"C:\\reports\\test_join.xlsx"
    writer = ReportXlsxSaver()
    writer.set_data(customer='ОАО «АМИГЭ»',
                    obj_name='Дополнительные инженерно-геологические изыскания на объекте: «Поисково-оценочная скважина № 5 Русановского лицензионного участка» в рамках договора по объекту «Бурение пяти инженерно-геологических скважин на шельфе Российской Федерации в 2021 году»,',
                    test_title='ВЕДОМОСТЬ РЕЗУЛЬТАТОВ ОПРЕДЕЛЕНИЯ ШТОРМОВОЙ РАЗЖИЖАЕМОСТИ ГРУНТОВ МЕТОДОМ ЦИКЛИЧЕСКИХ ТРЕХОСНЫХ СЖАТИЙ\n С РЕГУЛИРУЕМОЙ НАГРУЗКОЙ (ГОСТ 56353-2015, ASTM D5311/ASTM D5311M-13)',
                    date='14.03.2022',
                    tests_data={'22-63': {'skv': '1', 'depth': 22.1, 'type': 'Я просто камушек',
                                          'params': {"σ'1, кПа": 17,
                                                     "σ'3, кПа": 17,
                                                     "τ, кПа": 29,
                                                     "PPRmax, д.е.": 0.671,
                                                     "εmax, д.е.": 0.083,
                                                     "Nfail, ед.": "-",
                                                     "Результат испытания": "дин. неуст."}}},
                    additional_data=["Частота воздействия = 0,1 Гц",
                                     "Расчетная высота волны, м: 11,5",
                                     "Расчетное число циклов: 49371",
                                     "Плотность воды, кН/м3: 10"])
    writer.save("C:\\reports\\result.xlsx")
